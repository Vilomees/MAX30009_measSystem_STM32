#!/usr/bin/env python3
"""
MAX30009 BioZ Measurement Control Panel

PyQt5 GUI for controlling the MAX30009 bioimpedance AFE via STM32WB55 over UART.

Tabs:
  Overview          — serial port selection and live impedance dashboard
  BIA               — single-frequency continuous measurement
  BIS               — full spectroscopy scan (freq indices 0-59, 100 I&Q pairs each)
  Calibration Builder — builds a complex calibration JSON from 0Ω baseline + reference resistor CSVs
  Terminal / Log    — raw UART log and manual command entry

Measurement flow (BIS):
  1. GUI sends configuration commands (I, G, H, D, A, W, C, E) over UART to STM32.
  2. STM32 runs the scan and streams lines: SCAN_START → RAW,... × N → FREQ_DONE,... → SCAN_END.
  3. ScanWorker (background QThread) parses lines and emits Qt signals to update UI.
  4. Raw ADC rows (real, imag, flags per sample) are stored in memory and exported to CSV.
  5. For calibrated mode: two sequential scans (0Ω baseline then DUT) are compared in Excel.
"""

import sys
import os
import json
import csv
import math
import time
import datetime
import threading
import statistics
import re
import serial
import serial.tools.list_ports

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget,
    QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QLabel, QPushButton, QComboBox, QSpinBox, QDoubleSpinBox,
    QProgressBar, QTextEdit, QFileDialog, QGroupBox,
    QSplitter, QFrame, QSizePolicy, QStatusBar, QLineEdit,
    QCheckBox, QScrollArea, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox
)
from PyQt5.QtCore import (
    Qt, QTimer, pyqtSignal, QObject, QThread, QSize, QPointF, QMargins
)
from PyQt5.QtGui import (
    QFont, QColor, QPalette, QPainter, QPen, QBrush,
    QLinearGradient, QPolygonF, QFontDatabase
)
from PyQt5.QtChart import (
    QChart, QChartView, QLineSeries, QScatterSeries,
    QValueAxis, QLogValueAxis
)


# ──────────────────────────────────────────────────────────
# MAX30009 register-mapped constants (must match max30009_scan.py and firmware)
# ──────────────────────────────────────────────────────────
# Stimulus current options in µA, indexed by BIOZ_IDAC register field (0..15)
CURRENTS_UA = [
    0.016, 0.032, 0.080, 0.160, 0.320, 0.640,
    1.6,   3.2,   6.4,   12.8,  32.0,  64.0,
    128.0, 256.0, 640.0, 1280.0
]
CURRENTS_A = [c * 1e-6 for c in CURRENTS_UA]   # same values in amperes
GAINS = [1.0, 2.0, 5.0, 10.0]                  # receive channel PGA gain options (index 0..3)
# 60 frequency slots defined by MAX30009 PLL divider table; index matches firmware freq_idx
FREQ_HZ = [
    808960, 722944, 649216, 581632, 499712, 468992,
    420864, 377856, 338944, 304128, 272896, 249856,
    245248, 220160, 199936, 176896, 158976, 143106,
    131072, 114944,  99968,  93056,  82944,  82048,
     75008,  66944,  60032,  54016,  49984,  43008,
     41024,  38976,  35008,  30976,  28032,  24992,
     23008,  20000,  18016,  16000,  15008,  14016,
     13008,  12000,  11008,  10000,   9008,   8000,
      7008,   6000,   5000,   4000,   2000,   1000,
       500,    250,    125,     64,     32,     16
]
# Legacy OSR-based scaling constants (used only when the old K-table path is called).
# The active impedance formula uses IMP_DENOM_CONST instead (see below).
K_BY_ADC_OSR = {
    128:  77000,
    256:  154000,
    512:  110437,
    1024: 220874,
}

def get_adc_osr(freq_idx):
    """Return ADC oversampling ratio used by the firmware for a given frequency index.
    The OSR (and thus noise floor) changes at the OSR boundary indices 30, 51, 52."""
    if freq_idx <= 30:   return 128
    elif freq_idx <= 51: return 256
    elif freq_idx == 52: return 512
    else:                return 1024

def get_K(freq_idx, gain_idx=0, stim_idx=0):
    """Backward-compatible helper returning counts-per-ohm scaling.

    The MAX30009 impedance conversion follows the datasheet current-stimulus
    equation and is independent of OSR/frequency. freq_idx is kept only for
    compatibility with older call sites.
    """
    _ = freq_idx
    gain = GAINS[gain_idx]
    current_uA = CURRENTS_UA[stim_idx]
    return IMP_DENOM_CONST * gain * (current_uA * 1e-6)

def format_current(idx):
    c = CURRENTS_UA[idx]
    if c < 1:     return f"{c*1000:.0f} nA"
    elif c < 1000: return f"{c:.1f} µA"
    else:          return f"{c/1000:.2f} mA"

def format_freq(hz):
    if hz >= 1000000: return f"{hz/1000000:.3f} MHz"
    elif hz >= 1000:  return f"{hz/1000:.1f} kHz"
    else:             return f"{hz} Hz"


def fmt3(val):
    """Format number with up to 3 decimal places, trimming trailing zeros.
    Examples: 1.2345 -> "1.235", 1.0 -> "1", 1.500 -> "1.5", -0.0001 -> "0"."""
    if val is None:
        return ""
    try:
        v = float(val)
    except (TypeError, ValueError):
        return str(val)
    if math.isnan(v) or math.isinf(v):
        return str(v)
    s = f"{v:.3f}"
    if '.' in s:
        s = s.rstrip('0').rstrip('.')
    if s in ('-0', '-'):
        s = '0'
    return s


def fmt3_pct(val):
    """Format as percentage with up to 3 decimal places and '%' suffix."""
    if val is None or val == '':
        return ""
    try:
        v = float(val)
    except (TypeError, ValueError):
        return str(val)
    return fmt3(v) + " %"


# ──────────────────────────────────────────────────────────
# Impedance formula:  |Z| = MAG_ADC / (524288 · G · (2/π) · I_A)
#
# Derivation: the MAX30009 20-bit sigma-delta ADC full scale = 2^19 = 524288 counts.
# The demodulated sine amplitude is scaled by 2/π relative to the ADC full scale.
# Dividing raw ADC magnitude by (524288 · G · (2/π) · I[A]) gives ohms.
# ──────────────────────────────────────────────────────────
IMP_DENOM_CONST = 524288.0 * (2.0 / math.pi)  # ≈ 333772.6  (gain- and current-independent part)

def impedance_from_adc(real_adc, imag_adc, I_A, G):
    """|Z| = sqrt(real^2 + imag^2) / (524288 * G * (2/pi) * I)"""
    mag = math.sqrt(real_adc * real_adc + imag_adc * imag_adc)
    return mag / (IMP_DENOM_CONST * G * I_A)

def component_from_adc(adc_value, G, I_uA):
    """Single ADC component -> ohms using datasheet scaling.

    R and X must be calculated from their own ADC components, not from the
    already scaled K value multiplied again by gain/current.
    Formula: component_ohm = adc / (524288 * (2/pi) * G * I[A])
    """
    I_A = float(I_uA) * 1e-6
    G = float(G)
    if I_A == 0 or G == 0:
        return math.nan
    return float(adc_value) / (IMP_DENOM_CONST * G * I_A)

def parse_nominal_ohm(text):
    """Parseerib tekstist nominaaltakistuse oomides. Toetab '2p001K', '1k', '982.4', '1.5M' formaate."""
    if text is None:
        return None
    s = str(text).strip().replace(',', '.').replace('p', '.')
    if not s:
        return None
    m = re.match(r'^([0-9]*\.?[0-9]+)\s*([kKmMrRΩohm]*)', s)
    if not m:
        return None
    try:
        val = float(m.group(1))
    except ValueError:
        return None
    suffix = m.group(2).lower()
    if suffix.startswith('k'):
        val *= 1e3
    elif suffix.startswith('m'):
        val *= 1e6
    return val


# ──────────────────────────────────────────────────────────
# Color theme (dark, Task Manager style)
# ──────────────────────────────────────────────────────────
COLORS = {
    'bg':           '#f6f8fa',
    'bg2':          '#ffffff',
    'bg3':          '#eaeef2',
    'border':       '#d0d7de',
    'text':         '#1f2328',
    'text_dim':     '#656d76',
    'accent':       '#0969da',
    'accent2':      '#0969da',
    'green':        '#1a7f37',
    'yellow':       '#9a6700',
    'red':          '#cf222e',
    'cyan':         '#0969da',
    'purple':       '#8250df',
    'orange':       '#bc4c00',
}

STYLE_MAIN = f"""
QMainWindow, QWidget {{
    background-color: {COLORS['bg']};
    color: {COLORS['text']};
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 12px;
}}
QTabWidget::pane {{
    border: 1px solid {COLORS['border']};
    background: {COLORS['bg2']};
}}
QTabBar::tab {{
    background: {COLORS['bg3']};
    color: {COLORS['text_dim']};
    padding: 8px 20px;
    border: 1px solid {COLORS['border']};
    border-bottom: none;
    margin-right: 2px;
    font-size: 12px;
}}
QTabBar::tab:selected {{
    background: {COLORS['bg2']};
    color: {COLORS['text']};
    border-bottom: 2px solid {COLORS['accent2']};
}}
QTabBar::tab:hover {{
    background: {COLORS['bg2']};
    color: {COLORS['text']};
}}
QGroupBox {{
    border: 1px solid {COLORS['border']};
    border-radius: 4px;
    margin-top: 8px;
    padding-top: 8px;
    color: {COLORS['text_dim']};
    font-size: 11px;
    font-weight: bold;
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 8px;
    padding: 0 4px;
    color: {COLORS['accent2']};
    text-transform: uppercase;
    letter-spacing: 1px;
}}
QPushButton {{
    background: {COLORS['bg3']};
    color: {COLORS['text']};
    border: 1px solid {COLORS['border']};
    border-radius: 4px;
    padding: 6px 16px;
    font-family: 'Consolas', monospace;
    font-size: 12px;
}}
QPushButton:hover {{
    background: {COLORS['accent']};
    border-color: {COLORS['accent2']};
    color: white;
}}
QPushButton:pressed {{
    background: #1158c7;
}}
QPushButton:disabled {{
    background: {COLORS['bg']};
    color: {COLORS['text_dim']};
    border-color: {COLORS['border']};
}}
QPushButton#btn_start {{
    background: #dafbe1;
    border-color: {COLORS['green']};
    color: {COLORS['green']};
    font-weight: bold;
}}
QPushButton#btn_start:hover {{
    background: {COLORS['green']};
    color: {COLORS['bg']};
}}
QPushButton#btn_stop {{
    background: #ffebe9;
    border-color: {COLORS['red']};
    color: {COLORS['red']};
    font-weight: bold;
}}
QPushButton#btn_stop:hover {{
    background: {COLORS['red']};
    color: white;
}}
QComboBox, QSpinBox, QDoubleSpinBox, QLineEdit {{
    background: {COLORS['bg3']};
    color: {COLORS['text']};
    border: 1px solid {COLORS['border']};
    border-radius: 3px;
    padding: 4px 8px;
    font-family: 'Consolas', monospace;
    selection-background-color: {COLORS['accent']};
}}
QComboBox::drop-down {{
    border: none;
    width: 20px;
}}
QComboBox QAbstractItemView {{
    background: {COLORS['bg3']};
    color: {COLORS['text']};
    border: 1px solid {COLORS['border']};
    selection-background-color: {COLORS['accent']};
}}
QProgressBar {{
    background: {COLORS['bg3']};
    border: 1px solid {COLORS['border']};
    border-radius: 3px;
    height: 16px;
    text-align: center;
    color: {COLORS['text']};
    font-size: 11px;
}}
QProgressBar::chunk {{
    background: {COLORS['accent2']};
    border-radius: 2px;
}}
QTextEdit {{
    background: {COLORS['bg']};
    color: {COLORS['cyan']};
    border: 1px solid {COLORS['border']};
    border-radius: 3px;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 11px;
    selection-background-color: {COLORS['accent']};
}}
QTableWidget {{
    background: {COLORS['bg']};
    color: {COLORS['text']};
    border: 1px solid {COLORS['border']};
    gridline-color: {COLORS['border']};
    selection-background-color: {COLORS['accent']};
    font-family: 'Consolas', monospace;
    font-size: 11px;
}}
QTableWidget::item {{
    padding: 4px 8px;
    border: none;
}}
QHeaderView::section {{
    background: {COLORS['bg3']};
    color: {COLORS['accent2']};
    border: 1px solid {COLORS['border']};
    padding: 4px 8px;
    font-size: 11px;
    font-weight: bold;
    text-transform: uppercase;
    letter-spacing: 1px;
}}
QScrollBar:vertical {{
    background: {COLORS['bg']};
    width: 8px;
    border: none;
}}
QScrollBar::handle:vertical {{
    background: {COLORS['border']};
    border-radius: 4px;
    min-height: 20px;
}}
QScrollBar::handle:vertical:hover {{
    background: {COLORS['text_dim']};
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0px;
}}
QStatusBar {{
    background: {COLORS['bg3']};
    color: {COLORS['text_dim']};
    border-top: 1px solid {COLORS['border']};
    font-size: 11px;
}}
QLabel#metric_big {{
    color: {COLORS['accent2']};
    font-size: 28px;
    font-weight: bold;
    font-family: 'Consolas', monospace;
}}
QLabel#metric_unit {{
    color: {COLORS['text_dim']};
    font-size: 13px;
}}
QLabel#metric_label {{
    color: {COLORS['text_dim']};
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 1px;
}}
QLabel#status_ok {{
    color: {COLORS['green']};
    font-weight: bold;
}}
QLabel#status_err {{
    color: {COLORS['red']};
    font-weight: bold;
}}
QFrame#separator {{
    background: {COLORS['border']};
    max-height: 1px;
}}
QCheckBox {{
    color: {COLORS['text']};
    spacing: 6px;
}}
QCheckBox::indicator {{
    width: 14px;
    height: 14px;
    border: 1px solid {COLORS['border']};
    border-radius: 2px;
    background: {COLORS['bg3']};
}}
QCheckBox::indicator:checked {{
    background: {COLORS['accent2']};
    border-color: {COLORS['accent2']};
}}
"""


# ──────────────────────────────────────────────────────────
# Scrolling real-time mini-chart (Task-Manager style)
# ──────────────────────────────────────────────────────────
class MiniPlot(QWidget):
    """Fixed-width ring-buffer plot: add_point() appends, oldest point is dropped when full."""
    def __init__(self, color=COLORS['accent2'], label='', parent=None):
        super().__init__(parent)
        self.color = QColor(color)
        self.label = label
        self.data = []
        self.max_points = 120
        self.y_min = 0
        self.y_max = 1
        self.setMinimumHeight(80)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def add_point(self, value):
        self.data.append(value)
        if len(self.data) > self.max_points:
            self.data.pop(0)
        if self.data:
            self.y_min = min(self.data) * 0.9
            self.y_max = max(self.data) * 1.1
            if self.y_min == self.y_max:
                self.y_max = self.y_min + 1
        self.update()

    def clear(self):
        self.data = []
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()

        # Taust
        painter.fillRect(0, 0, w, h, QColor(COLORS['bg']))

        # Ruudustik
        painter.setPen(QPen(QColor(COLORS['border']), 1, Qt.DotLine))
        for i in range(1, 4):
            y = int(h * i / 4)
            painter.drawLine(0, y, w, y)

        if len(self.data) < 2:
            # Label
            painter.setPen(QColor(COLORS['text_dim']))
            painter.setFont(QFont('Consolas', 9))
            painter.drawText(4, h - 4, self.label)
            return

        # Gradient fill
        fill_color = QColor(self.color)
        fill_color.setAlpha(40)
        grad = QLinearGradient(0, 0, 0, h)
        grad.setColorAt(0, fill_color)
        grad.setColorAt(1, QColor(0, 0, 0, 0))

        def to_x(i): return int(i * w / (self.max_points - 1))
        def to_y(v):
            ratio = (v - self.y_min) / (self.y_max - self.y_min)
            return int(h - ratio * (h - 4) - 2)

        # Polygon for area fill
        points = [QPolygonF()]
        poly = points[0]
        from PyQt5.QtCore import QPointF
        offset = self.max_points - len(self.data)
        poly.append(QPointF(to_x(offset), h))
        for i, v in enumerate(self.data):
            poly.append(QPointF(to_x(i + offset), to_y(v)))
        poly.append(QPointF(to_x(offset + len(self.data) - 1), h))

        painter.setBrush(QBrush(grad))
        painter.setPen(Qt.NoPen)
        painter.drawPolygon(poly)

        # Joon
        painter.setPen(QPen(self.color, 1.5))
        for i in range(1, len(self.data)):
            x1 = to_x(i - 1 + offset)
            y1 = to_y(self.data[i-1])
            x2 = to_x(i + offset)
            y2 = to_y(self.data[i])
            painter.drawLine(x1, y1, x2, y2)

        # Label
        painter.setPen(QColor(COLORS['text_dim']))
        painter.setFont(QFont('Consolas', 9))
        painter.drawText(4, h - 4, self.label)

        # Latest value label
        if self.data:
            painter.setPen(self.color)
            val_str = fmt3(self.data[-1])
            painter.drawText(w - 60, 14, val_str)


# ──────────────────────────────────────────────────────────
# Large impedance readout widget
# ──────────────────────────────────────────────────────────
class MetricCard(QWidget):
    def __init__(self, label, unit, color=COLORS['accent2'], parent=None):
        super().__init__(parent)
        self.setObjectName("metric_card")
        layout = QVBoxLayout(self)
        layout.setSpacing(2)
        layout.setContentsMargins(12, 10, 12, 10)

        self.lbl_label = QLabel(label.upper())
        self.lbl_label.setObjectName("metric_label")
        self.lbl_label.setAlignment(Qt.AlignLeft)

        row = QHBoxLayout()
        self.lbl_value = QLabel("—")
        self.lbl_value.setObjectName("metric_big")
        self.lbl_value.setStyleSheet(f"color: {color};")
        self.lbl_value.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        self.lbl_unit = QLabel(unit)
        self.lbl_unit.setObjectName("metric_unit")
        self.lbl_unit.setAlignment(Qt.AlignLeft | Qt.AlignBottom)
        self.lbl_unit.setStyleSheet("padding-bottom: 4px;")

        row.addWidget(self.lbl_value)
        row.addWidget(self.lbl_unit)
        row.addStretch()

        layout.addWidget(self.lbl_label)
        layout.addLayout(row)

        self.setStyleSheet(f"""
            QWidget#metric_card {{
                background: {COLORS['bg2']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
            }}
        """)

    def set_value(self, val):
        if val is None:
            self.lbl_value.setText("—")
            return
        try:
            v = float(val)
        except (TypeError, ValueError):
            self.lbl_value.setText(str(val))
            return
        av = abs(v)
        if av >= 1_000_000:
            self.lbl_value.setText(fmt3(v / 1_000_000) + "M")
        elif av >= 1_000:
            self.lbl_value.setText(fmt3(v / 1_000) + "k")
        else:
            self.lbl_value.setText(fmt3(v))


# ──────────────────────────────────────────────────────────
# Background worker: opens UART, drives STM32, parses streaming data
# Runs in a dedicated QThread so the GUI never blocks.
# ──────────────────────────────────────────────────────────
class ScanWorker(QObject):
    progress = pyqtSignal(int, int)      # (completed_freqs, total_freqs) — drives progress bar
    log = pyqtSignal(str)                # text line for the Terminal tab
    data_point = pyqtSignal(dict)        # every raw ADC row as it arrives
    freq_done = pyqtSignal(int, int, float, float)  # (freq_idx, n_samples, R_ohm, X_ohm) after each frequency
    finished = pyqtSignal(bool, str)     # (success, message) when scan ends or aborts
    single_result = pyqtSignal(float, float, float, float, float, float)  # R, X, |Z|, phase, avg_real, avg_imag
    raw_pair = pyqtSignal(int, int, int)  # (real_adc, imag_adc, flags) — streamed live to LiveADCWindow
    flags_result = pyqtSignal(int)        # OR of all flags bits in the last measurement batch

    def __init__(self):
        super().__init__()
        self._stop = False
        self.port = ''
        self.baud = 1000000
        self.stim_idx = 12
        self.gain_idx = 0
        self.freq_start = 0
        self.freq_end = 52
        self.component = 'test'
        self.mode = 'scan'
        self.single_freq = 30
        # Advanced Settings
        self.ahpf = 15       # AHPF bypass (res open + cap short)
        self.dhpf = 0        # Digi HPF bypass
        self.dlpf = 0        # Digi LPF bypass
        self.dcrestore = 0   # DC restore disabled
        self.extcap = 0      # External cap disabled
        self.amprange = 3    # Amp Range High
        self.ampbw = 3       # Amp BW High
        self.log_mode = 'averaged'  # 'averaged' | 'raw'
        self.single_raw_buf = []    # accumulates all raw I&Q pairs during single measurement

    def stop(self):
        self._stop = True

    def run_scan(self):
        """Full spectroscopy sweep: sends R freq_start freq_end to STM32, collects all raw pairs.

        Protocol (UART, ASCII):
          PC → STM32:  I<n>  G<n>  H<n>  D<n n>  A<n>  W<n>  C<n>  E<n>  R <start> <end>
          STM32 → PC:  SCAN_START
                       RAW,<fi>,<hz>,<sn>,<real>,<imag>,<flags>  (×100 per frequency)
                       FREQ_DONE,<fi>,<n>
                       ...repeated for each frequency...
                       SCAN_END
        """
        self._stop = False
        self.raw_rows_buf = []  # raw ADC rows buffer for GUI export
        try:
            ser = serial.Serial(self.port, self.baud, timeout=10)
        except Exception as e:
            self.finished.emit(False, f"Connection error: {e}")
            return

        time.sleep(0.5)
        ser.reset_input_buffer()
        ser.reset_output_buffer()
        time.sleep(0.1)

        ser.write(b'\n')
        time.sleep(0.3)
        ser.reset_input_buffer()

        self.log.emit(f"[UART] I{self.stim_idx} ({format_current(self.stim_idx)})")
        ser.write(f'I{self.stim_idx}\n'.encode())
        time.sleep(0.3)
        while ser.in_waiting:
            l = ser.readline().decode('ascii', errors='ignore').strip()
            if l: self.log.emit(f"[STM32] {l}")

        self.log.emit(f"[UART] G{self.gain_idx} ({GAINS[self.gain_idx]:.0f}x)")
        ser.write(f'G{self.gain_idx}\n'.encode())
        time.sleep(0.3)
        while ser.in_waiting:
            l = ser.readline().decode('ascii', errors='ignore').strip()
            if l: self.log.emit(f"[STM32] {l}")

        # ── Advanced Settings ──
        def send_cmd(cmd_str, desc):
            self.log.emit(f"[UART] {cmd_str}  ({desc})")
            ser.write((cmd_str + '\n').encode())
            time.sleep(0.25)
            while ser.in_waiting:
                l = ser.readline().decode('ascii', errors='ignore').strip()
                if l: self.log.emit(f"[STM32] {l}")

        send_cmd(f'H{self.ahpf}',                       f'AHPF idx={self.ahpf}')
        send_cmd(f'D{self.dhpf} {self.dlpf}',            f'DigFilter DHPF={self.dhpf} DLPF={self.dlpf}')
        send_cmd(f'A{self.amprange}',                    f'AmpRange={self.amprange}')
        send_cmd(f'W{self.ampbw}',                       f'AmpBW={self.ampbw}')
        send_cmd(f'C{self.dcrestore}',                   f'DCRestore={self.dcrestore}')
        send_cmd(f'E{self.extcap}',                      f'ExtCap={self.extcap}')

        cmd = f'R {self.freq_start} {self.freq_end}\n'
        self.log.emit(f"[UART] {cmd.strip()}")
        ser.write(cmd.encode())
        time.sleep(0.3)

        rows = []
        freq_data = {}  # freq_idx -> [real_adc, ...]
        n_total = self.freq_end - self.freq_start + 1
        scan_started = False
        scan_finished = False

        while not scan_finished and not self._stop:
            try:
                raw = ser.readline()
                if not raw:
                    if not scan_started:
                        self.log.emit("WARNING: Timeout, waiting...")
                    continue
                line = raw.decode('ascii', errors='ignore').strip()
                if not line:
                    continue
            except Exception as e:
                self.log.emit(f"Serial error: {e}")
                break

            if line.startswith('SCAN_START'):
                scan_started = True
                self.log.emit(f"✓ {line}")

            elif (line.startswith('RAW,') or line.startswith('PAIR,')) and scan_started:
                # One I&Q pair per line: RAW,<fi>,<freq_hz>,<sample_nr>,<real_adc>,<imag_adc>,<flags>
                parts = line.split(',')
                if len(parts) >= 7:
                    try:
                        fi = int(parts[1])
                        fhz = int(parts[2])
                        sn = int(parts[3])
                        real_adc = int(parts[4])
                        imag_adc = int(parts[5])
                        flags = int(parts[6])
                        row = {'freq_idx': fi, 'freq_hz': fhz, 'sample_nr': sn,
                               'real_adc': real_adc, 'imag_adc': imag_adc, 'flags': flags}
                        rows.append(row)
                        self.raw_rows_buf.append(row)
                        if fi not in freq_data:
                            freq_data[fi] = {'real': [], 'imag': []}
                        freq_data[fi]['real'].append(real_adc)
                        freq_data[fi]['imag'].append(imag_adc)
                        self.data_point.emit(row)
                    except ValueError:
                        pass

            elif line.startswith('FREQ_DONE'):
                # All samples for this frequency received; compute and emit averaged R/X
                parts = line.split(',')
                if len(parts) >= 3:
                    try:
                        fi = int(parts[1])
                        ns = int(parts[2])
                        done = len(freq_data)
                        self.progress.emit(done, n_total)
                        if fi in freq_data and freq_data[fi]['real']:
                            avg_r = sum(freq_data[fi]['real']) / len(freq_data[fi]['real'])
                            avg_i = sum(freq_data[fi]['imag']) / len(freq_data[fi]['imag'])
                            G = GAINS[self.gain_idx]
                            I_uA = CURRENTS_UA[self.stim_idx]
                            R_ohm = component_from_adc(avg_r, G, I_uA)
                            X_ohm = component_from_adc(avg_i, G, I_uA)
                            self.freq_done.emit(fi, ns, R_ohm, X_ohm)
                    except (ValueError, IndexError):
                        pass

            elif line == 'SCAN_END':
                scan_finished = True
                self.log.emit("✓ SCAN_END")

        ser.close()

        if self._stop:
            self.finished.emit(False, "Stopped by user")
            return

        if rows:
            self.finished.emit(True, f"{len(rows)} raw ADC rows collected")
        else:
            self.finished.emit(False, "No data received")

    def run_single(self):
        """Continuous measurement at one frequency — repeats R <fi> <fi> until stopped.

        Each iteration triggers one complete 100-sample batch at the selected frequency.
        R/X/|Z|/phase are averaged within the batch and emitted via single_result.
        The raw pairs buffer (single_raw_buf) accumulates across all iterations for export.
        """
        self._stop = False
        self.single_raw_buf = []  # reset at the start of each measurement run
        try:
            ser = serial.Serial(self.port, self.baud, timeout=3)
        except Exception as e:
            self.finished.emit(False, f"Connection error: {e}")
            return

        time.sleep(0.3)
        ser.reset_input_buffer()
        ser.write(b'\n')
        time.sleep(0.2)
        ser.reset_input_buffer()

        def send_cmd(cmd_str, desc=""):
            self.log.emit(f"[UART] {cmd_str}  ({desc})" if desc else f"[UART] {cmd_str}")
            ser.write((cmd_str + '\n').encode())
            time.sleep(0.25)
            while ser.in_waiting:
                l = ser.readline().decode('ascii', errors='ignore').strip()
                if l: self.log.emit(f"[STM32] {l}")

        # ── Advanced filter/amp settings before main measurement ──
        send_cmd(f'H{self.ahpf}',             f'AHPF idx={self.ahpf}')
        send_cmd(f'D{self.dhpf} {self.dlpf}', f'DigFilter DHPF={self.dhpf} DLPF={self.dlpf}')
        send_cmd(f'A{self.amprange}',         f'AmpRange={self.amprange}')
        send_cmd(f'W{self.ampbw}',            f'AmpBW={self.ampbw}')
        send_cmd(f'C{self.dcrestore}',        f'DCRestore={self.dcrestore}')
        send_cmd(f'E{self.extcap}',           f'ExtCap={self.extcap}')

        # ── Stim + Gain + Freq ──
        send_cmd(f'I{self.stim_idx}',  format_current(self.stim_idx))
        send_cmd(f'G{self.gain_idx}',  f'{GAINS[self.gain_idx]:.0f}x')

        freq_hz = FREQ_HZ[self.single_freq] if self.single_freq < len(FREQ_HZ) else 0
        send_cmd(f'F{self.single_freq}', format_freq(freq_hz))
        time.sleep(0.3)
        ser.reset_input_buffer()

        meas_count = 0
        # Unbuffered Z/R/X/Phase accumulator for averaging
        z_vals, r_vals, x_vals, ph_vals = [], [], [], []

        # Continuous measurement: repeat R command at one frequency (same protocol as scan,
        # but a single frequency only). More reliable than the S command.
        fi = self.single_freq

        while not self._stop:
            cmd = f'R {fi} {fi}\n'
            self.log.emit(f"[UART] {cmd.strip()}")
            ser.write(cmd.encode())

            real_vals, imag_vals, flags_vals = [], [], []
            scan_started = False
            deadline = time.time() + 15

            while time.time() < deadline and not self._stop:
                try:
                    raw = ser.readline()
                except Exception as e:
                    self.log.emit(f"Serial error: {e}")
                    break
                if not raw:
                    continue
                line = raw.decode('ascii', errors='ignore').strip()
                if not line:
                    continue

                if line.startswith('SCAN_START'):
                    scan_started = True
                    self.log.emit(f"[STM32] {line}")

                elif (line.startswith('RAW,') or line.startswith('PAIR,')) and scan_started:
                    parts = line.split(',')
                    if len(parts) >= 7:
                        try:
                            rv = int(parts[4])
                            iv = int(parts[5])
                            fl = int(parts[6])
                            real_vals.append(rv)
                            imag_vals.append(iv)
                            flags_vals.append(fl)
                            self.single_raw_buf.append({
                                'meas_nr': meas_count + 1,
                                'sample_nr': len(real_vals) - 1,
                                'real_adc': rv,
                                'imag_adc': iv,
                                'flags': fl,
                            })
                            self.raw_pair.emit(rv, iv, fl)  # kohe, ilma bufferita
                        except ValueError:
                            pass

                elif line == 'SCAN_END':
                    break

                elif line.startswith('DBG,') or line.startswith('RD:'):
                    pass

                else:
                    if not scan_started:
                        self.log.emit(f"[STM32] {line}")

            if real_vals and not self._stop:
                G = GAINS[self.gain_idx]
                I_uA = CURRENTS_UA[self.stim_idx]
                avg_r = sum(real_vals) / len(real_vals)
                avg_i = sum(imag_vals) / len(imag_vals)
                R = component_from_adc(avg_r, G, I_uA)
                X = component_from_adc(avg_i, G, I_uA)
                mag = math.sqrt(R*R + X*X)
                phase = math.degrees(math.atan2(X, R))
                meas_count += 1
                combined_flags = 0
                for f in flags_vals:
                    combined_flags |= f
                self.single_result.emit(R, X, mag, phase, avg_r, avg_i)
                self.flags_result.emit(combined_flags)
                if self.log_mode == 'raw':
                    for sn, (rv, iv) in enumerate(zip(real_vals, imag_vals)):
                        self.log.emit(
                            f"[#{meas_count:3d}][{sn:3d}] real={rv}  imag={iv}"
                        )
                    self.log.emit(
                        f"[#{meas_count:3d}] --- R={fmt3(R)}Ω  X={fmt3(X)}Ω  "
                        f"|Z|={fmt3(mag)}Ω  φ={fmt3(phase)}°  (n={len(real_vals)})"
                    )
                else:
                    self.log.emit(
                        f"[#{meas_count:3d}] R={fmt3(R)}Ω  X={fmt3(X)}Ω  "
                        f"|Z|={fmt3(mag)}Ω  φ={fmt3(phase)}°  "
                        f"avg_real={fmt3(avg_r)}  avg_imag={fmt3(avg_i)}  (n={len(real_vals)})"
                    )
            elif not self._stop and not real_vals:
                self.log.emit("⚠ No data received — check the connection")
                time.sleep(1)

        ser.write(b'P\n')
        time.sleep(0.2)
        ser.close()
        self.finished.emit(True, f"Stopped — {meas_count} measurements completed")


# ──────────────────────────────────────────────────────────
# Tab 1 — Overview: serial connection + live metric cards + mini-plots
# ──────────────────────────────────────────────────────────
class DashboardTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # ── Top row: connection controls ──
        conn_group = QGroupBox("Connection")
        conn_layout = QHBoxLayout(conn_group)

        self.cb_port = QComboBox()
        self.cb_port.setMinimumWidth(180)
        self._refresh_ports()

        self.btn_refresh = QPushButton("↺")
        self.btn_refresh.setFixedWidth(32)
        self.btn_refresh.setToolTip("Refresh Port List")
        self.btn_refresh.clicked.connect(self._refresh_ports)

        self.cb_baud = QComboBox()
        for b in ['115200', '230400', '460800', '921600', '1000000']:
            self.cb_baud.addItem(b)
        self.cb_baud.setCurrentText('1000000')

        self.btn_connect = QPushButton("Connect")
        self.btn_connect.clicked.connect(self._test_connect)

        self.lbl_conn_status = QLabel("● No Connection")
        self.lbl_conn_status.setObjectName("status_err")

        conn_layout.addWidget(QLabel("Port:"))
        conn_layout.addWidget(self.cb_port)
        conn_layout.addWidget(self.btn_refresh)
        conn_layout.addWidget(QLabel("  Baud:"))
        conn_layout.addWidget(self.cb_baud)
        conn_layout.addWidget(self.btn_connect)
        conn_layout.addStretch()
        conn_layout.addWidget(self.lbl_conn_status)

        layout.addWidget(conn_group)

        # ── Metric cards ──
        metrics_layout = QHBoxLayout()
        self.card_R = MetricCard("Resistance R", "Ω", COLORS['accent2'])
        self.card_X = MetricCard("Reactance X", "Ω", COLORS['purple'])
        self.card_Z = MetricCard("Impedance |Z|", "Ω", COLORS['green'])
        self.card_P = MetricCard("Phase φ", "°", COLORS['orange'])

        metrics_layout.addWidget(self.card_R)
        metrics_layout.addWidget(self.card_X)
        metrics_layout.addWidget(self.card_Z)
        metrics_layout.addWidget(self.card_P)
        layout.addLayout(metrics_layout)

        # ── Reaalajas graafikud ──
        plots_group = QGroupBox("Real-time — latest values")
        plots_layout = QGridLayout(plots_group)

        self.plot_R = MiniPlot(COLORS['accent2'], '|Z| Ω')
        self.plot_phase = MiniPlot(COLORS['orange'], 'Phase °')

        plots_layout.addWidget(self.plot_R, 0, 0)
        plots_layout.addWidget(self.plot_phase, 0, 1)
        layout.addWidget(plots_group)

        # ── Frequencye info ──
        info_group = QGroupBox("Configuration")
        info_layout = QGridLayout(info_group)
        info_layout.setColumnStretch(1, 1)
        info_layout.setColumnStretch(3, 1)

        self.lbl_freq_info = QLabel("—")
        self.lbl_current_info = QLabel("—")
        self.lbl_gain_info = QLabel("—")
        self.lbl_samples_info = QLabel("—")

        for lbl in [self.lbl_freq_info, self.lbl_current_info,
                    self.lbl_gain_info, self.lbl_samples_info]:
            lbl.setStyleSheet(f"color: {COLORS['text']}; font-size: 12px;")

        info_layout.addWidget(QLabel("Frequency:"), 0, 0)
        info_layout.addWidget(self.lbl_freq_info, 0, 1)
        info_layout.addWidget(QLabel("Current:"), 0, 2)
        info_layout.addWidget(self.lbl_current_info, 0, 3)
        info_layout.addWidget(QLabel("Gain:"), 1, 0)
        info_layout.addWidget(self.lbl_gain_info, 1, 1)
        info_layout.addWidget(QLabel("Samples:"), 1, 2)
        info_layout.addWidget(self.lbl_samples_info, 1, 3)
        layout.addWidget(info_group)

    def _refresh_ports(self):
        self.cb_port.clear()
        ports = serial.tools.list_ports.comports()
        for p in ports:
            self.cb_port.addItem(p.device)
        if not ports:
            self.cb_port.addItem("(No ports available)")

    def _test_connect(self):
        port = self.cb_port.currentText()
        baud = int(self.cb_baud.currentText())
        try:
            s = serial.Serial(port, baud, timeout=1)
            s.write(b'?\n')
            time.sleep(0.3)
            resp = ''
            while s.in_waiting:
                resp += s.readline().decode('ascii', errors='ignore').strip() + ' '
            s.close()
            self.lbl_conn_status.setText("● Connected")
            self.lbl_conn_status.setObjectName("status_ok")
            self.lbl_conn_status.setStyleSheet(f"color: {COLORS['green']}; font-weight: bold;")
        except Exception as e:
            self.lbl_conn_status.setText("● Connection Error")
            self.lbl_conn_status.setObjectName("status_err")
            self.lbl_conn_status.setStyleSheet(f"color: {COLORS['red']}; font-weight: bold;")

    def update_single_result(self, R, X, Z, phase):
        self.card_R.set_value(R)
        self.card_X.set_value(X)
        self.card_Z.set_value(Z)
        self.card_P.set_value(phase)
        self.plot_R.add_point(Z)
        self.plot_phase.add_point(phase)

    def get_port(self):
        return self.cb_port.currentText()

    def get_baud(self):
        return int(self.cb_baud.currentText())


# ──────────────────────────────────────────────────────────
# Tab 2 — BIA (single-frequency measurement)
# ──────────────────────────────────────────────────────────
class SingleMeasTab(QWidget):
    start_requested = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.history_export_rows = []
        self.calibration_data = None
        self.calibration_path = None
        self._live_window = None
        self._current_worker = None
        self._last_flags = 0  # viimase batch'i flags
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        # ═════════════════════════════════════════════
        # TOP ROW: left = main settings, right = advanced settings
        # ═════════════════════════════════════════════
        settings_row = QHBoxLayout()
        settings_row.setSpacing(8)

        # ── Left: main settings ──
        cfg_group = QGroupBox("Measurement Settings")
        cfg_form = QFormLayout(cfg_group)
        cfg_form.setLabelAlignment(Qt.AlignRight)
        cfg_form.setContentsMargins(8, 8, 8, 8)
        cfg_form.setVerticalSpacing(4)
        cfg_form.setHorizontalSpacing(6)

        self.le_component = QLineEdit("")
        self.le_component.setPlaceholderText("e.g. 982.4")
        cfg_form.addRow("Reference resistor (Ω):", self.le_component)

        self.cb_current = QComboBox()
        for i in range(16):
            self.cb_current.addItem(f"[{i:2d}]  {format_current(i)}")
        self.cb_current.setCurrentIndex(12)
        cfg_form.addRow("Stimulation current:", self.cb_current)

        self.cb_gain = QComboBox()
        for i, g in enumerate(GAINS):
            self.cb_gain.addItem(f"[{i}]  {g:.0f}x")
        cfg_form.addRow("Gain:", self.cb_gain)

        self.cb_freq = QComboBox()
        for i, hz in enumerate(FREQ_HZ):
            osr = get_adc_osr(i)
            self.cb_freq.addItem(f"[{i:2d}]  {format_freq(hz)}  (OSR={osr})")
        self.cb_freq.setCurrentIndex(30)
        cfg_form.addRow("Frequency:", self.cb_freq)

        self.cb_log_mode = QComboBox()
        self.cb_log_mode.addItem("Averaged (R, X, |Z|, φ)")
        self.cb_log_mode.addItem("Raw (ADC I&Q pairs)")
        self.cb_log_mode.setCurrentIndex(0)
        cfg_form.addRow("Log mode:", self.cb_log_mode)

        settings_row.addWidget(cfg_group, 1)

        # ── Right: advanced settings (same as BIS tab) ──
        adv_group = QGroupBox("Advanced Settings")
        adv_form = QFormLayout(adv_group)
        adv_form.setLabelAlignment(Qt.AlignRight)
        adv_form.setContentsMargins(8, 8, 8, 8)
        adv_form.setVerticalSpacing(4)
        adv_form.setHorizontalSpacing(6)

        # Analoog HPF
        self.cb_ahpf = QComboBox()
        ahpf_opts = [
            "[0]  100 Hz", "[1]  200 Hz", "[2]  500 Hz", "[3]  1 kHz",
            "[4]  2 kHz", "[5]  5 kHz", "[6]  10 kHz",
            "[7]  Bypass (res. open)",
            "[8]  42.4 MΩ + cap short", "[9]  21.2 MΩ + cap short",
            "[10] 8.4 MΩ + cap short", "[11] 4.2 MΩ + cap short",
            "[12] 2.2 MΩ + cap short", "[13] 848 kΩ + cap short",
            "[14] 848 kΩ + cap short",
            "[15] Full bypass (res. open + cap short)",
        ]
        for o in ahpf_opts:
            self.cb_ahpf.addItem(o)
        self.cb_ahpf.setCurrentIndex(15)
        adv_form.addRow("Analog HPF:", self.cb_ahpf)

        # Digitaalfilter HPF
        self.cb_dhpf = QComboBox()
        for idx, lbl in enumerate(["[0]  Bypass", "[1]  0.00025 × SR", "[2]  0.002 × SR", "[3]  0.002 × SR"]):
            self.cb_dhpf.addItem(lbl)
        self.cb_dhpf.setCurrentIndex(0)
        adv_form.addRow("Digital HPF:", self.cb_dhpf)

        # Digitaalfilter LPF
        self.cb_dlpf = QComboBox()
        for idx, lbl in enumerate(["[0]  Bypass", "[1]  0.005 × SR", "[2]  0.02 × SR", "[3]  0.08 × SR",
                                    "[4]  0.25 × SR", "[5]  0.25 × SR", "[6]  0.25 × SR", "[7]  0.25 × SR"]):
            self.cb_dlpf.addItem(lbl)
        self.cb_dlpf.setCurrentIndex(0)
        adv_form.addRow("Digital LPF:", self.cb_dlpf)

        # DC Restore
        self.cb_dcrestore = QComboBox()
        self.cb_dcrestore.addItem("[0]  Off (default)")
        self.cb_dcrestore.addItem("[1]  On (10 MΩ feedback)")
        self.cb_dcrestore.setCurrentIndex(0)
        adv_form.addRow("DC Restore:", self.cb_dcrestore)

        # Ext Cap
        self.cb_extcap = QComboBox()
        self.cb_extcap.addItem("[0]  Off (default)")
        self.cb_extcap.addItem("[1]  On (CEXT DRVXC-DRVSJ)")
        self.cb_extcap.setCurrentIndex(0)
        adv_form.addRow("Ext Cap:", self.cb_extcap)

        # Amp Range (IDRV_RGE)
        self.cb_amprange = QComboBox()
        for idx, lbl in enumerate(["[0]  Low", "[1]  Medium-Low", "[2]  Medium-High", "[3]  High"]):
            self.cb_amprange.addItem(lbl)
        self.cb_amprange.setCurrentIndex(3)
        adv_form.addRow("Amp Range:", self.cb_amprange)

        # Amp BW
        self.cb_ampbw = QComboBox()
        for idx, lbl in enumerate(["[0]  Low", "[1]  Medium-Low", "[2]  Medium-High", "[3]  High"]):
            self.cb_ampbw.addItem(lbl)
        self.cb_ampbw.setCurrentIndex(3)
        adv_form.addRow("Amp BW:", self.cb_ampbw)

        settings_row.addWidget(adv_group, 1)
        layout.addLayout(settings_row)

        # ═════════════════════════════════════════════
        # KALIBRATSIOONI RIDA
        # ═════════════════════════════════════════════
        cal_group = QGroupBox("Calibration (optional)")
        cal_hbox = QHBoxLayout(cal_group)
        cal_hbox.setContentsMargins(8, 6, 8, 6)
        cal_hbox.setSpacing(6)

        self.le_cal_path = QLineEdit()
        self.le_cal_path.setReadOnly(True)
        self.le_cal_path.setPlaceholderText("No calibration file loaded — results shown uncalibrated")
        self.btn_cal_load = QPushButton("Load JSON…")
        self.btn_cal_load.setFixedWidth(90)
        self.btn_cal_load.clicked.connect(self._browse_calibration)
        self.btn_cal_clear = QPushButton("Clear")
        self.btn_cal_clear.setFixedWidth(55)
        self.btn_cal_clear.clicked.connect(self._clear_calibration)
        self.lbl_cal_info = QLabel("")
        self.lbl_cal_info.setStyleSheet(f"color: {COLORS.get('text_dim','#888')}; font-size: 10px;")

        cal_hbox.addWidget(QLabel("Calibration JSON:"))
        cal_hbox.addWidget(self.le_cal_path, 1)
        cal_hbox.addWidget(self.btn_cal_load)
        cal_hbox.addWidget(self.btn_cal_clear)
        cal_hbox.addWidget(self.lbl_cal_info)
        layout.addWidget(cal_group)

        # ═════════════════════════════════════════════
        # BUTTONS (narrow row)
        # ═════════════════════════════════════════════
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.btn_start = QPushButton("▶  Start Measurement")
        self.btn_start.setObjectName("btn_start")
        self.btn_start.setMinimumHeight(32)
        self.btn_start.clicked.connect(self._on_start)

        self.btn_stop = QPushButton("■  Stop")
        self.btn_stop.setObjectName("btn_stop")
        self.btn_stop.setMinimumHeight(32)
        self.btn_stop.setEnabled(False)

        self.btn_live = QPushButton("📈  Live ADC")
        self.btn_live.setMinimumHeight(32)
        self.btn_live.setToolTip("Open live ADC I&Q plot window")
        self.btn_live.clicked.connect(self._open_live_window)

        btn_row.addWidget(self.btn_start)
        btn_row.addWidget(self.btn_stop)
        btn_row.addWidget(self.btn_live)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        # ═════════════════════════════════════════════
        # BOTTOM ROW: left = result cards, right = measurement history
        # ═════════════════════════════════════════════
        bottom_row = QHBoxLayout()
        bottom_row.setSpacing(8)

        # Result: 2x2 grid, kitsas
        res_group = QGroupBox("Result")
        res_layout = QGridLayout(res_group)
        res_layout.setContentsMargins(8, 8, 8, 8)
        res_layout.setSpacing(6)

        self.card_R = MetricCard("Resistance R", "Ω", COLORS['accent2'])
        self.card_X = MetricCard("Reactance X", "Ω", COLORS['purple'])
        self.card_Z = MetricCard("Impedance |Z|", "Ω", COLORS['green'])
        self.card_P = MetricCard("Phase φ", "°", COLORS['orange'])

        res_layout.addWidget(self.card_R, 0, 0)
        res_layout.addWidget(self.card_X, 0, 1)
        res_layout.addWidget(self.card_Z, 1, 0)
        res_layout.addWidget(self.card_P, 1, 1)

        # Flags indikaatorid
        FLAG_DEFS = [
            ('IDRV_OVR', 0, '#ff5555', 'Drive current out of range'),
            ('CODE_OVR', 1, '#ffb86c', 'ADC positive overflow'),
            ('CODE_UND', 2, '#ffb86c', 'ADC negative underflow'),
            ('LEADS_OFF', 3, '#ff79c6', 'Lead-off detected'),
            ('ERROR',    4, '#ff5555', 'General error'),
            ('MARKER',   5, '#8be9fd', 'Marker bit'),
        ]
        flags_row = QHBoxLayout()
        flags_row.setSpacing(4)
        self._flag_labels = {}
        for name, bit, color, tip in FLAG_DEFS:
            lbl = QLabel(name)
            lbl.setToolTip(tip)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setFixedHeight(18)
            lbl.setStyleSheet(
                f"font-size: 9px; font-weight: bold; border-radius: 3px; padding: 0 4px;"
                f"background: #2a2a3a; color: #555566;"
            )
            self._flag_labels[bit] = (lbl, color)
            flags_row.addWidget(lbl)
        res_layout.addLayout(flags_row, 2, 0, 1, 2)
        bottom_row.addWidget(res_group, 1)

        # History table (side by side with result)
        hist_group = QGroupBox("Measurement History")
        hist_layout = QVBoxLayout(hist_group)
        hist_layout.setContentsMargins(8, 8, 8, 8)
        hist_layout.setSpacing(4)
        self.tbl_history = QTableWidget(0, 12)
        self.tbl_history.setHorizontalHeaderLabels(
            ['Time', 'Reference (Ω)', 'Freq',
             'R (Ω)', 'X (Ω)', '|Z| (Ω)', 'ADC real', 'ADC imag',
             'R_cal (Ω)', '|Z|_cal (Ω)', 'φ_cal (°)', 'Flags'])
        hist_header = self.tbl_history.horizontalHeader()
        hist_header.setSectionResizeMode(QHeaderView.Interactive)
        hist_header.setDefaultAlignment(Qt.AlignCenter)
        self.tbl_history.verticalHeader().setDefaultSectionSize(24)
        self.tbl_history.setStyleSheet(
            self.tbl_history.styleSheet() +
            "QHeaderView::section { font-size: 10px; padding: 2px 4px; }"
            "QTableWidget { font-size: 10px; }"
            "QTableWidget::item { padding: 2px 4px; }"
        )
        for col, width in enumerate([70, 118, 78, 82, 82, 82, 86, 86, 86, 86, 72, 110]):
            self.tbl_history.setColumnWidth(col, width)
        self.tbl_history.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_history.setEditTriggers(QTableWidget.NoEditTriggers)
        hist_layout.addWidget(self.tbl_history)

        hist_actions = QHBoxLayout()
        btn_export = QPushButton("Export History CSV")
        btn_export.clicked.connect(self.export_history_csv)
        hist_actions.addWidget(btn_export)

        self.chk_save_raw = QCheckBox("Save raw pairs")
        self.chk_save_raw.setToolTip(
            "If checked, all raw ADC I&Q pairs are collected in memory.\n"
            "Use 'Export Raw CSV' to save them to disk.")
        hist_actions.addWidget(self.chk_save_raw)

        self.btn_export_raw = QPushButton("Export Raw CSV")
        self.btn_export_raw.setToolTip("Save all collected raw ADC I&Q pairs to a CSV file")
        self.btn_export_raw.clicked.connect(self.export_raw_csv)
        hist_actions.addWidget(self.btn_export_raw)

        btn_clear = QPushButton("Clear History")
        btn_clear.clicked.connect(self.clear_history)
        hist_actions.addWidget(btn_clear)
        hist_layout.addLayout(hist_actions)
        bottom_row.addWidget(hist_group, 2)

        layout.addLayout(bottom_row)

    def _on_start(self):
        cfg = {
            'mode': 'single',
            'component': self.le_component.text().strip() or 'test',
            'stim_idx': self.cb_current.currentIndex(),
            'gain_idx': self.cb_gain.currentIndex(),
            'single_freq': self.cb_freq.currentIndex(),
            'ahpf':      self.cb_ahpf.currentIndex(),
            'dhpf':      self.cb_dhpf.currentIndex(),
            'dlpf':      self.cb_dlpf.currentIndex(),
            'dcrestore': self.cb_dcrestore.currentIndex(),
            'extcap':    self.cb_extcap.currentIndex(),
            'amprange':  self.cb_amprange.currentIndex(),
            'ampbw':     self.cb_ampbw.currentIndex(),
            'log_mode':  'raw' if self.cb_log_mode.currentIndex() == 1 else 'averaged',
        }
        self.start_requested.emit(cfg)

    def set_running(self, running):
        self.btn_start.setEnabled(not running)
        self.btn_stop.setEnabled(running)

    def set_worker(self, worker):
        self._current_worker = worker
        # If checkbox unchecked, do not accumulate — clear previous buffer
        if not self.chk_save_raw.isChecked():
            worker.single_raw_buf.clear()
        # If window is already open, connect immediately
        if self._live_window is not None and self._live_window.isVisible():
            try:
                worker.raw_pair.connect(self._live_window.add_pair)
            except Exception:
                pass

    def _open_live_window(self):
        if self._live_window is None or not self._live_window.isVisible():
            self._live_window = LiveADCWindow()
            # Connect current worker if measurement is running
            if self._current_worker is not None:
                try:
                    self._current_worker.raw_pair.connect(self._live_window.add_pair)
                except Exception:
                    pass
            self._live_window.show()
        else:
            self._live_window.raise_()
            self._live_window.activateWindow()

    def get_live_window(self):
        return self._live_window

    def update_flags(self, combined_flags):
        """Uuenda flag indikaatoreid (OR'd flags viimase batch'i kohta)."""
        self._last_flags = combined_flags
        for bit, (lbl, active_color) in self._flag_labels.items():
            if combined_flags & (1 << bit):
                lbl.setStyleSheet(
                    f"font-size: 9px; font-weight: bold; border-radius: 3px; padding: 0 4px;"
                    f"background: {active_color}; color: #11111b;"
                )
            else:
                lbl.setStyleSheet(
                    f"font-size: 9px; font-weight: bold; border-radius: 3px; padding: 0 4px;"
                    f"background: #2a2a3a; color: #555566;"
                )

    def update_result(self, R, X, Z, phase, component, freq_idx, avg_r=0.0, avg_i=0.0):
        self.card_R.set_value(R)
        self.card_X.set_value(X)
        self.card_Z.set_value(Z)
        self.card_P.set_value(phase)

        # Kalibratsiooni rakendamine (kui JSON on laetud)
        cal_result = self.apply_calibration(freq_idx, avg_r, avg_i)
        if cal_result is not None:
            R_cal, X_cal, Z_cal, phase_cal = cal_result
            self.card_R.set_value(R_cal)
            self.card_X.set_value(X_cal)
            self.card_Z.set_value(Z_cal)
            self.card_P.set_value(phase_cal)

        row = self.tbl_history.rowCount()
        self.tbl_history.insertRow(row)
        ts = datetime.datetime.now().strftime('%H:%M:%S')
        freq_str = format_freq(FREQ_HZ[freq_idx]) if freq_idx < len(FREQ_HZ) else '?'

        base_vals = [ts, component, freq_str,
                     fmt3(R), fmt3(X), fmt3(Z),
                     fmt3(avg_r), fmt3(avg_i)]
        if cal_result is not None:
            R_cal, X_cal, Z_cal, phase_cal = cal_result
            cal_vals = [fmt3(R_cal), fmt3(Z_cal), fmt3(phase_cal)]
        else:
            cal_vals = ['—', '—', '—']

        for col, val in enumerate(base_vals + cal_vals):
            item = QTableWidgetItem(val)
            item.setTextAlignment(Qt.AlignCenter)
            if col >= 8 and cal_result is not None:
                item.setForeground(QColor('#4fc3f7'))
            self.tbl_history.setItem(row, col, item)

        # Flags veerg (col 11)
        fl = self._last_flags
        FLAG_NAMES = {0:'IDRV_OVR', 1:'CODE_OVR', 2:'CODE_UND', 3:'LEADS_OFF', 4:'ERROR', 5:'MARKER'}
        if fl == 0:
            flags_str = 'OK'
            flags_color = '#50fa7b'
        else:
            active = [FLAG_NAMES[b] for b in range(6) if fl & (1 << b)]
            flags_str = ' '.join(active)
            flags_color = '#ff5555'
        fl_item = QTableWidgetItem(flags_str)
        fl_item.setTextAlignment(Qt.AlignCenter)
        fl_item.setForeground(QColor(flags_color))
        self.tbl_history.setItem(row, 11, fl_item)
        self.tbl_history.scrollToBottom()

        export_row = {
            "measurement_index": row + 1,
            "time": ts,
            "component_or_reference": component,
            "frequency_index": freq_idx,
            "frequency_hz": FREQ_HZ[freq_idx] if freq_idx < len(FREQ_HZ) else '',
            "frequency_label": freq_str,
            "stimulation_current_uA": CURRENTS_UA[self.cb_current.currentIndex()],
            "gain": GAINS[self.cb_gain.currentIndex()],
            "adc_real": avg_r,
            "adc_imag": avg_i,
            "resistance_ohm": R,
            "reactance_ohm": X,
            "impedance_ohm": Z,
            "phase_deg": phase,
        }
        if cal_result is not None:
            R_cal, X_cal, Z_cal, phase_cal = cal_result
            export_row.update({
                "R_cal_ohm": R_cal,
                "X_cal_ohm": X_cal,
                "Z_cal_ohm": Z_cal,
                "phase_cal_deg": phase_cal,
            })
        self.history_export_rows.append(export_row)

    def clear_history(self):
        self.tbl_history.setRowCount(0)
        self.history_export_rows.clear()
        if self._current_worker is not None:
            self._current_worker.single_raw_buf.clear()

    def export_history_csv(self):
        if not self.history_export_rows:
            QMessageBox.warning(self, "Export Error", "Measurement history is empty.")
            return

        default_name = f"single_measurement_history_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Measurement History",
            default_name,
            "CSV Files (*.csv)"
        )
        if not path:
            return

        fieldnames = [
            "measurement_index",
            "time",
            "component_or_reference",
            "frequency_index",
            "frequency_hz",
            "frequency_label",
            "stimulation_current_uA",
            "gain",
            "adc_real",
            "adc_imag",
            "resistance_ohm",
            "reactance_ohm",
            "impedance_ohm",
            "phase_deg",
            "R_cal_ohm",
            "X_cal_ohm",
            "Z_cal_ohm",
            "phase_cal_deg",
        ]
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(self.history_export_rows)
            QMessageBox.information(self, "Export Complete", f"Measurement history exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export measurement history:\n{e}")

    def export_raw_csv(self):
        """Export all raw ADC I&Q pairs collected during single measurement runs."""
        if self._current_worker is None or not self._current_worker.single_raw_buf:
            QMessageBox.warning(self, "Export Error",
                "Raw paarid puuduvad.\n"
                "Make sure 'Save raw pairs' is checked and a measurement has been run.")
            return

        if not self.chk_save_raw.isChecked():
            reply = QMessageBox.question(
                self, "Save raw pairs",
                "'Save raw pairs' is not checked — export existing data anyway?",
                QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes:
                return

        buf = self._current_worker.single_raw_buf
        default_name = f"single_raw_pairs_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Raw ADC Pairs", default_name, "CSV Files (*.csv)")
        if not path:
            return

        stim_idx = self.cb_current.currentIndex()
        gain_idx = self.cb_gain.currentIndex()
        freq_idx = self.cb_freq.currentIndex()
        header_lines = [
            "# MAX30009 BIA — Raw ADC I&Q paarid",
            f"# timestamp: {datetime.datetime.now().isoformat()}",
            f"# component: {self.le_component.text().strip() or 'test'}",
            f"# freq_idx: {freq_idx}",
            f"# freq_hz: {FREQ_HZ[freq_idx] if freq_idx < len(FREQ_HZ) else '?'}",
            f"# stim_idx: {stim_idx}",
            f"# stim_uA: {CURRENTS_UA[stim_idx]}",
            f"# gain_idx: {gain_idx}",
            f"# gain_x: {GAINS[gain_idx]:.0f}",
            f"# ahpf_idx: {self.cb_ahpf.currentIndex()}",
            f"# dhpf_idx: {self.cb_dhpf.currentIndex()}",
            f"# dlpf_idx: {self.cb_dlpf.currentIndex()}",
            f"# dcrestore: {self.cb_dcrestore.currentIndex()}",
            f"# extcap: {self.cb_extcap.currentIndex()}",
            f"# amprange: {self.cb_amprange.currentIndex()}",
            f"# ampbw: {self.cb_ampbw.currentIndex()}",
            "#",
        ]

        fieldnames = ['meas_nr', 'sample_nr', 'real_adc', 'imag_adc', 'flags']
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                for ln in header_lines:
                    f.write(ln + "\n")
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(buf)
            QMessageBox.information(self, "Export Complete",
                f"{len(buf)} raw pairs exported:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Export failed:\n{e}")

    def _browse_calibration(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Calibration JSON", self.calibration_path or "",
            "JSON Files (*.json)")
        if not path:
            return
        try:
            with open(path, encoding='utf-8') as f:
                cal = json.load(f)
            if 'calibration' not in cal or 'config' not in cal:
                raise ValueError("Missing 'calibration' or 'config' key")
            self.calibration_data = cal
            self.calibration_path = path
            n = len(cal['calibration'])
            cfg = cal['config']
            self.le_cal_path.setText(os.path.basename(path))
            self.lbl_cal_info.setText(
                f"{n} freqs | stim={cfg.get('stim_uA','?')}µA "
                f"gain={cfg.get('gain_x','?')}x | ref={cal.get('reference_resistor_ohm','?')}Ω")
        except Exception as e:
            QMessageBox.critical(self, "Calibration Load Error", str(e))
            self.calibration_data = None
            self.calibration_path = None

    def _clear_calibration(self):
        self.calibration_data = None
        self.calibration_path = None
        self.le_cal_path.clear()
        self.lbl_cal_info.setText("")

    def apply_calibration(self, freq_idx, avg_r, avg_i):
        """Apply complex calibration correction to raw ADC averages.

        The calibration JSON stores a per-frequency complex factor k = Z_nominal / Z_measured
        computed during the Calibration Builder step.  Applying it:
            Z_corrected = k(f) * (ADC_net(f) / K)
        where ADC_net = current_measurement - baseline (0Ω offset subtracted).

        Returns (R_cal, X_cal, |Z|_cal, phase_cal_deg) or None if no calibration loaded.
        """
        import cmath
        if not self.calibration_data:
            return None
        cal_entry = self.calibration_data['calibration'].get(str(freq_idx))
        if cal_entry is None:
            return None
        K = self.calibration_data['config']['K_value']
        b_r = cal_entry['baseline_adc_real']
        b_i = cal_entry['baseline_adc_imag']
        k = complex(cal_entry['k_real'], cal_entry['k_imag'])
        adc_net = complex(avg_r - b_r, avg_i - b_i)
        z_cal = k * (adc_net / K)
        mag = abs(z_cal)
        phase = math.degrees(cmath.phase(z_cal)) if mag > 0 else 0.0
        return z_cal.real, z_cal.imag, mag, phase


# ──────────────────────────────────────────────────────────
# Floating window: live I & Q ADC time-series during a single measurement
# ──────────────────────────────────────────────────────────
class LiveADCWindow(QWidget):
    """Separate window that connects to ScanWorker.raw_pair signal and plots each sample."""

    MAX_POINTS = 500  # mitu punkti graafikul

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Live ADC — I & Q")
        self.setMinimumSize(820, 480)
        self._sample_idx = 0
        self._last_real = 0
        self._last_imag = 0
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        # ── Top bar: channel selection ──
        top = QHBoxLayout()
        top.setSpacing(8)

        lbl_ch = QLabel("Channel:")
        lbl_ch.setStyleSheet("color: #a6adc8; font-size: 11px;")
        self.btn_ch_both = QPushButton("I + Q")
        self.btn_ch_i    = QPushButton("I only")
        self.btn_ch_q    = QPushButton("Q only")
        for btn in (self.btn_ch_both, self.btn_ch_i, self.btn_ch_q):
            btn.setCheckable(True)
            btn.setFixedHeight(24)
            btn.setFixedWidth(70)
        self.btn_ch_both.setChecked(True)
        self.btn_ch_both.clicked.connect(lambda: self._set_channel('both'))
        self.btn_ch_i.clicked.connect(lambda: self._set_channel('I'))
        self.btn_ch_q.clicked.connect(lambda: self._set_channel('Q'))
        self._channel = 'both'

        top.addWidget(lbl_ch)
        top.addWidget(self.btn_ch_both)
        top.addWidget(self.btn_ch_i)
        top.addWidget(self.btn_ch_q)
        top.addStretch()
        layout.addLayout(top)

        # ── Graafik ──
        self._chart = QChart()
        self._chart.setTitle("")
        self._chart.setBackgroundBrush(QBrush(QColor('#ffffff')))
        self._chart.setTitleBrush(QBrush(QColor('#1f2328')))
        self._chart.legend().setLabelColor(QColor('#1f2328'))
        self._chart.setMargins(QMargins(4, 4, 4, 4))
        self._chart.layout().setContentsMargins(0, 0, 0, 0)

        self._series_i = QLineSeries()
        self._series_i.setName("I  (real)")
        pen_i = QPen(QColor('#89b4fa'))
        pen_i.setWidth(1)
        self._series_i.setPen(pen_i)

        self._series_q = QLineSeries()
        self._series_q.setName("Q  (imag)")
        pen_q = QPen(QColor('#f38ba8'))
        pen_q.setWidth(1)
        self._series_q.setPen(pen_q)

        self._chart.addSeries(self._series_i)
        self._chart.addSeries(self._series_q)

        self._axis_x = QValueAxis()
        self._axis_x.setLabelFormat('%d')
        self._axis_x.setTitleText("Sample")
        self._axis_x.setLabelsColor(QColor('#656d76'))
        self._axis_x.setTitleBrush(QBrush(QColor('#656d76')))
        self._axis_x.setGridLineColor(QColor('#d0d7de'))
        self._axis_x.setRange(0, self.MAX_POINTS)

        self._axis_y = QValueAxis()
        self._axis_y.setLabelFormat('%d')
        self._axis_y.setTitleText("ADC counts")
        self._axis_y.setLabelsColor(QColor('#656d76'))
        self._axis_y.setTitleBrush(QBrush(QColor('#656d76')))
        self._axis_y.setGridLineColor(QColor('#d0d7de'))
        self._axis_y.setRange(-40000, 40000)

        self._chart.addAxis(self._axis_x, Qt.AlignBottom)
        self._chart.addAxis(self._axis_y, Qt.AlignLeft)
        self._series_i.attachAxis(self._axis_x)
        self._series_i.attachAxis(self._axis_y)
        self._series_q.attachAxis(self._axis_x)
        self._series_q.attachAxis(self._axis_y)

        self._chart_view = QChartView(self._chart)
        self._chart_view.setRenderHint(QPainter.Antialiasing, False)
        self._chart_view.setStyleSheet("background: #ffffff;")
        layout.addWidget(self._chart_view, 1)

        # ── Bottom row: readout labels + controls ──
        bot = QHBoxLayout()
        bot.setSpacing(12)

        # Latest value readouts
        val_frame = QFrame()
        val_frame.setStyleSheet(f"background: {COLORS['bg2']}; border-radius: 4px; padding: 2px;")
        val_layout = QHBoxLayout(val_frame)
        val_layout.setContentsMargins(10, 4, 10, 4)
        val_layout.setSpacing(20)

        self.lbl_i = QLabel("I: —")
        self.lbl_q = QLabel("Q: —")
        self.lbl_n = QLabel("n: 0")
        for lbl, color in [(self.lbl_i, '#89b4fa'), (self.lbl_q, '#f38ba8'), (self.lbl_n, '#a6adc8')]:
            lbl.setStyleSheet(f"color: {color}; font-family: Consolas; font-size: 12px; font-weight: bold;")
            val_layout.addWidget(lbl)

        bot.addWidget(val_frame)
        bot.addStretch()

        # ── Y zoom kontroll ──
        y_ctrl = QFrame()
        y_ctrl.setStyleSheet(f"background: {COLORS['bg2']}; border-radius: 4px;")
        y_layout = QHBoxLayout(y_ctrl)
        y_layout.setContentsMargins(8, 2, 8, 2)
        y_layout.setSpacing(6)

        lbl_center = QLabel("Center:")
        lbl_center.setStyleSheet("color: #a6adc8; font-size: 10px;")
        _sb_style = ("QSpinBox { background: #ffffff; color: #1f2328; border: 1px solid #d0d7de;"
                     " border-radius: 3px; padding: 1px 4px; }"
                     "QSpinBox::up-button, QSpinBox::down-button { width: 14px; }")
        self.sb_center = QSpinBox()
        self.sb_center.setRange(-2**23, 2**23)
        self.sb_center.setValue(0)
        self.sb_center.setSingleStep(1000)
        self.sb_center.setFixedWidth(90)
        self.sb_center.setToolTip("Y axis center (ADC counts)")
        self.sb_center.setStyleSheet(_sb_style)
        self.sb_center.valueChanged.connect(self._apply_y_range)

        lbl_span = QLabel("Span:")
        lbl_span.setStyleSheet("color: #a6adc8; font-size: 10px;")
        self.sb_span = QSpinBox()
        self.sb_span.setRange(10, 2**24)
        self.sb_span.setValue(80000)
        self.sb_span.setSingleStep(1000)
        self.sb_span.setFixedWidth(90)
        self.sb_span.setToolTip("Y axis total span (ADC counts) — smaller = more zoom")
        self.sb_span.setStyleSheet(_sb_style)
        self.sb_span.valueChanged.connect(self._apply_y_range)

        _btn_style_base = (
            f"QPushButton {{ background: {COLORS['bg3']}; color: {COLORS['text']};"
            f" border: 1px solid {COLORS['border']}; border-radius: 3px; padding: 3px 8px; }}"
            f"QPushButton:hover {{ background: {COLORS['accent']}; color: white; border-color: {COLORS['accent']}; }}"
        )

        btn_autoscale = QPushButton("Auto Y")
        btn_autoscale.setToolTip("Fit Y axis to all visible data (both channels)")
        btn_autoscale.setStyleSheet(_btn_style_base)
        btn_autoscale.clicked.connect(self._autoscale_y)

        btn_center_i = QPushButton("Center I")
        btn_center_i.setToolTip("Center Y on I channel median, keep current Span")
        btn_center_i.setStyleSheet(
            f"QPushButton {{ background: {COLORS['bg3']}; color: #0969da;"
            f" border: 1px solid #0969da; border-radius: 3px; padding: 3px 8px; }}"
            f"QPushButton:hover {{ background: #0969da; color: white; }}"
        )
        btn_center_i.clicked.connect(lambda: self._center_on_channel('I'))

        btn_center_q = QPushButton("Center Q")
        btn_center_q.setToolTip("Center Y on Q channel median, keep current Span")
        btn_center_q.setStyleSheet(
            f"QPushButton {{ background: {COLORS['bg3']}; color: #cf222e;"
            f" border: 1px solid #cf222e; border-radius: 3px; padding: 3px 8px; }}"
            f"QPushButton:hover {{ background: #cf222e; color: white; }}"
        )
        btn_center_q.clicked.connect(lambda: self._center_on_channel('Q'))

        btn_clear = QPushButton("Clear")
        btn_clear.setStyleSheet(_btn_style_base)
        btn_clear.clicked.connect(self.clear)

        y_layout.addWidget(lbl_center)
        y_layout.addWidget(self.sb_center)
        y_layout.addWidget(lbl_span)
        y_layout.addWidget(self.sb_span)
        y_layout.addWidget(btn_center_i)
        y_layout.addWidget(btn_center_q)
        y_layout.addWidget(btn_autoscale)
        y_layout.addWidget(btn_clear)

        bot.addWidget(y_ctrl)
        layout.addLayout(bot)

        # Buffer punktide jaoks
        self._buf_i = []
        self._buf_q = []
        self._y_min = -40000
        self._y_max = 40000

    def _set_channel(self, ch):
        self._channel = ch
        self.btn_ch_both.setChecked(ch == 'both')
        self.btn_ch_i.setChecked(ch == 'I')
        self.btn_ch_q.setChecked(ch == 'Q')
        self._series_i.setVisible(ch in ('both', 'I'))
        self._series_q.setVisible(ch in ('both', 'Q'))

    def _center_on_channel(self, ch):
        """Seadista Center kanali mediaanile, hoia praegune Span."""
        buf = self._buf_i if ch == 'I' else self._buf_q
        if not buf:
            return
        vals = [p.y() for p in buf]
        vals.sort()
        median = vals[len(vals) // 2]
        self.sb_center.blockSignals(True)
        self.sb_center.setValue(int(median))
        self.sb_center.blockSignals(False)
        self._apply_y_range()

    def _apply_y_range(self):
        center = self.sb_center.value()
        half = self.sb_span.value() / 2
        self._y_min = center - half
        self._y_max = center + half
        self._axis_y.setRange(self._y_min, self._y_max)

    def add_pair(self, real_adc, imag_adc, flags=0):
        """Kutsutakse iga saabunud I&Q paari jaoks — peab olema kiire."""
        x = self._sample_idx
        self._sample_idx += 1
        self._last_real = real_adc
        self._last_imag = imag_adc

        self._buf_i.append(QPointF(x, real_adc))
        self._buf_q.append(QPointF(x, imag_adc))

        if len(self._buf_i) > self.MAX_POINTS:
            self._buf_i.pop(0)
            self._buf_q.pop(0)

        if self._channel in ('both', 'I'):
            self._series_i.replace(self._buf_i)
        if self._channel in ('both', 'Q'):
            self._series_q.replace(self._buf_q)

        x_end = x + 1
        x_start = max(0, x_end - self.MAX_POINTS)
        self._axis_x.setRange(x_start, x_end)

        # Uuenda labelid
        self.lbl_i.setText(f"I: {real_adc:+d}")
        self.lbl_q.setText(f"Q: {imag_adc:+d}")
        self.lbl_n.setText(f"n: {self._sample_idx}")

    def clear(self):
        self._buf_i.clear()
        self._buf_q.clear()
        self._series_i.clear()
        self._series_q.clear()
        self._sample_idx = 0
        self._axis_x.setRange(0, self.MAX_POINTS)
        self._apply_y_range()
        self.lbl_i.setText("I: —")
        self.lbl_q.setText("Q: —")
        self.lbl_n.setText("n: 0")

    def _autoscale_y(self):
        bufs = []
        if self._channel in ('both', 'I') and self._buf_i:
            bufs += [p.y() for p in self._buf_i]
        if self._channel in ('both', 'Q') and self._buf_q:
            bufs += [p.y() for p in self._buf_q]
        if not bufs:
            return
        mn, mx = min(bufs), max(bufs)
        span = max(mx - mn, 10)
        margin = span * 0.15
        center = (mn + mx) / 2
        new_span = int(span + 2 * margin)
        self.sb_center.blockSignals(True)
        self.sb_span.blockSignals(True)
        self.sb_center.setValue(int(center))
        self.sb_span.setValue(new_span)
        self.sb_center.blockSignals(False)
        self.sb_span.blockSignals(False)
        self._apply_y_range()


# ──────────────────────────────────────────────────────────
# Tab: Calibration Builder
# ──────────────────────────────────────────────────────────
class CalibrationTab(QWidget):
    """Build a calibration JSON from two raw scan CSVs (0Ω baseline + reference resistor)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # ── Sisendid ──
        inp_group = QGroupBox("Input Files")
        inp_form = QFormLayout(inp_group)
        inp_form.setLabelAlignment(Qt.AlignRight)
        inp_form.setContentsMargins(10, 10, 10, 10)
        inp_form.setVerticalSpacing(6)

        # 0Ω baseline CSV
        row0 = QHBoxLayout()
        self.le_baseline = QLineEdit()
        self.le_baseline.setReadOnly(True)
        self.le_baseline.setPlaceholderText("Select 0Ω baseline raw CSV…")
        btn_b = QPushButton("Browse…")
        btn_b.setFixedWidth(80)
        btn_b.clicked.connect(lambda: self._browse_csv(self.le_baseline, "0Ω baseline"))
        row0.addWidget(self.le_baseline)
        row0.addWidget(btn_b)
        inp_form.addRow("0Ω baseline CSV:", row0)

        # Reference resistor CSV
        row1 = QHBoxLayout()
        self.le_refcsv = QLineEdit()
        self.le_refcsv.setReadOnly(True)
        self.le_refcsv.setPlaceholderText("Select reference resistor raw CSV…")
        btn_r = QPushButton("Browse…")
        btn_r.setFixedWidth(80)
        btn_r.clicked.connect(lambda: self._browse_csv(self.le_refcsv, "reference"))
        row1.addWidget(self.le_refcsv)
        row1.addWidget(btn_r)
        inp_form.addRow("Reference CSV:", row1)

        # Nominal resistance input
        self.dsb_nominal = QDoubleSpinBox()
        self.dsb_nominal.setRange(0.001, 1e6)
        self.dsb_nominal.setDecimals(4)
        self.dsb_nominal.setValue(100.0)
        self.dsb_nominal.setSuffix("  Ω")
        inp_form.addRow("Nominal value (HP34401A):", self.dsb_nominal)

        # Output JSON path
        row2 = QHBoxLayout()
        self.le_outpath = QLineEdit()
        self.le_outpath.setPlaceholderText("Output JSON path (auto-generated if empty)")
        btn_o = QPushButton("Save as…")
        btn_o.setFixedWidth(80)
        btn_o.clicked.connect(self._browse_output)
        row2.addWidget(self.le_outpath)
        row2.addWidget(btn_o)
        inp_form.addRow("Output JSON:", row2)

        layout.addWidget(inp_group)

        # ── Nupp ──
        btn_row = QHBoxLayout()
        self.btn_build = QPushButton("▶  Build Calibration JSON")
        self.btn_build.setObjectName("btn_start")
        self.btn_build.setMinimumHeight(34)
        self.btn_build.clicked.connect(self._build_calibration)
        btn_row.addWidget(self.btn_build)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        # ── Log / summary output ──
        self.te_log = QTextEdit()
        self.te_log.setReadOnly(True)
        self.te_log.setFont(QFont("Consolas", 9))
        self.te_log.setMinimumHeight(200)
        layout.addWidget(self.te_log, 1)

    def _browse_csv(self, line_edit, label):
        path, _ = QFileDialog.getOpenFileName(
            self, f"Select {label} CSV", "", "CSV Files (*.csv)")
        if path:
            line_edit.setText(path)

    def _browse_output(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Calibration JSON", "calibration.json", "JSON Files (*.json)")
        if path:
            self.le_outpath.setText(path)

    def _log(self, msg):
        self.te_log.append(msg)

    def _build_calibration(self):
        """Compute per-frequency complex calibration factor k from two raw scan CSVs.

        Inputs:
          baseline CSV — scan with 0Ω (open/shorted leads, measuring wire/fixture offset)
          reference CSV — scan with a known precision resistor (nominal_ohm)

        For each frequency index:
          ADC_net  = complex(ref_avg_real - base_avg_real,  ref_avg_imag - base_avg_imag)
          Z_meas   = ADC_net / K           (K = 524288 · G · (2/π) · I_A)
          k        = nominal_ohm / Z_meas  (complex correction factor)

        Output: calibration JSON with k_real, k_imag, k_abs, k_arg_deg per frequency.
        Apply later as:  Z_corrected = k · (ADC_DUT_net / K)
        """
        import cmath, statistics as stats
        ref_path = self.le_refcsv.text().strip()
        nominal = self.dsb_nominal.value()

        if not baseline_path or not ref_path:
            QMessageBox.warning(self, "Missing Input", "Select both CSV files first.")
            return

        self.te_log.clear()
        self._log("── Building calibration ──")
        self._log(f"  Baseline : {os.path.basename(baseline_path)}")
        self._log(f"  Reference: {os.path.basename(ref_path)}")
        self._log(f"  Nominal  : {nominal} Ω")

        def load_raw_csv(path):
            data = {}
            meta = {}
            with open(path, encoding='utf-8-sig') as f:
                for line in f:
                    line = line.rstrip('\r\n')
                    if line.startswith('#'):
                        if ':' in line:
                            k, v = line[1:].strip().split(':', 1)
                            meta[k.strip()] = v.strip()
                        continue
                    if line.startswith('freq_idx') or not line.strip():
                        continue
                    parts = line.split(',')
                    if len(parts) < 7:
                        continue
                    fi = int(parts[0])
                    fhz = int(parts[1]); osr = int(parts[2])
                    sn = int(parts[3]); r = int(parts[4])
                    i = int(parts[5]); fl = int(parts[6])
                    data.setdefault(fi, []).append((r, i, fl, fhz, osr, sn))
            return data, meta

        try:
            base_data, base_meta = load_raw_csv(baseline_path)
            ref_data, ref_meta = load_raw_csv(ref_path)
        except Exception as e:
            QMessageBox.critical(self, "CSV Load Error", str(e))
            self._log(f"✗ Load error: {e}")
            return

        # Kontroll
        for key in ['stim_idx', 'gain_idx']:
            bv = base_meta.get(key); rv = ref_meta.get(key)
            if bv != rv:
                self._log(f"  ⚠ {key}: baseline={bv}  ref={rv}")

        try:
            stim_idx = int(ref_meta['stim_idx'])
            stim_uA  = float(ref_meta['stim_uA'])
            gain_idx = int(ref_meta['gain_idx'])
            gain_x   = float(ref_meta['gain_x'])
            ahpf_idx = int(ref_meta.get('ahpf_idx', 0))
            dhpf_idx = int(ref_meta.get('dhpf_idx', 0))
            dlpf_idx = int(ref_meta.get('dlpf_idx', 0))
        except (KeyError, ValueError) as e:
            msg = (f"Could not parse CSV header: {e}\n\n"
                   f"These CSVs must be raw scan files exported from the BIS tab "
                   f"(they contain # header lines with stim_idx, gain_idx, stim_uA, gain_x).\n"
                   f"BIA CSV exports cannot be used here.")
            QMessageBox.critical(self, "Header Parse Error", msg)
            self._log(f"✗ {e}")
            return

        I_A = stim_uA * 1e-6
        G   = gain_x
        K   = 524288.0 * G * (2.0 / math.pi) * I_A
        self._log(f"  K = {K:.4f}  (stim={stim_uA}µA, gain={gain_x}x)")

        cal_table = {}
        for fi in sorted(base_data.keys()):
            if fi not in ref_data:
                self._log(f"  ⚠ F={fi}: missing in reference, skipped")
                continue
            b_r = stats.mean(s[0] for s in base_data[fi])
            b_i = stats.mean(s[1] for s in base_data[fi])
            r_r = stats.mean(s[0] for s in ref_data[fi])
            r_i = stats.mean(s[1] for s in ref_data[fi])

            adc_net = complex(r_r - b_r, r_i - b_i)
            z_meas  = adc_net / K
            if abs(z_meas) < 1e-9:
                self._log(f"  ⚠ F={fi}: Z_meas≈0, skipped")
                continue
            k = complex(nominal, 0) / z_meas
            std_r = stats.stdev(s[0] for s in ref_data[fi]) if len(ref_data[fi]) > 1 else 0
            std_i = stats.stdev(s[1] for s in ref_data[fi]) if len(ref_data[fi]) > 1 else 0

            cal_table[fi] = {
                'freq_hz': ref_data[fi][0][3],
                'osr': ref_data[fi][0][4],
                'k_real': k.real,
                'k_imag': k.imag,
                'k_abs': abs(k),
                'k_arg_deg': math.degrees(cmath.phase(k)),
                'z_meas_real': z_meas.real,
                'z_meas_imag': z_meas.imag,
                'ref_adc_std_real': std_r,
                'ref_adc_std_imag': std_i,
                'baseline_adc_real': b_r,
                'baseline_adc_imag': b_i,
            }

        cal_json = {
            'schema_version': 1,
            'description': 'MAX30009 complex calibration — apply as: Z_corr(f) = k(f) * (Z_DUT(f) - Z_baseline(f))',
            'reference_resistor_ohm': nominal,
            'reference_source': os.path.basename(ref_path),
            'baseline_source': os.path.basename(baseline_path),
            'reference_timestamp': ref_meta.get('timestamp', '-'),
            'baseline_timestamp': base_meta.get('timestamp', '-'),
            'config': {
                'stim_idx': stim_idx, 'stim_uA': stim_uA,
                'gain_idx': gain_idx, 'gain_x': gain_x,
                'ahpf_idx': ahpf_idx, 'dhpf_idx': dhpf_idx, 'dlpf_idx': dlpf_idx,
                'K_formula': '524288 * G * (2/pi) * I_A',
                'K_value': K,
            },
            'calibration': {str(fi): v for fi, v in cal_table.items()},
        }

        out_path = self.le_outpath.text().strip()
        if not out_path:
            out_path = os.path.join(
                os.path.dirname(ref_path),
                f"calibration_{nominal:g}ohm_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            )
            self.le_outpath.setText(out_path)

        try:
            with open(out_path, 'w') as f:
                json.dump(cal_json, f, indent=2)
        except Exception as e:
            QMessageBox.critical(self, "Save Error", str(e))
            self._log(f"✗ Save error: {e}")
            return

        self._log(f"\n✓ Calibration built: {len(cal_table)} frequencies")
        self._log(f"  Saved: {out_path}\n")
        self._log(f"{'F':>3}  {'Hz':>8}  {'|k|':>9}  {'arg(k)°':>10}")
        self._log("─" * 38)
        for fi_str, v in sorted(cal_json['calibration'].items(), key=lambda x: int(x[0])):
            fi = int(fi_str)
            if fi % 8 == 0 or fi in (30, 31, 52, 53):
                self._log(f"{fi:>3}  {v['freq_hz']:>8}  {v['k_abs']:>9.4f}  {v['k_arg_deg']:>+10.4f}°")

        QMessageBox.information(self, "Done",
            f"Calibration JSON saved to:\n{out_path}\n\n"
            f"{len(cal_table)} frequencies calibrated.")


# ──────────────────────────────────────────────────────────
# Tab 3 — BIS (Bioimpedance Spectroscopy): full frequency sweep
# ──────────────────────────────────────────────────────────
class ScanTab(QWidget):
    """Orchestrates the two-phase calibrated measurement workflow:
      Phase 1 (baseline): scan with 0Ω → store raw_rows + save CSV
      Phase 2 (DUT):      scan with resistor → compute offset-corrected |Z| → write to Excel
    The calib_mode / calib_phase state machine is driven by _on_finished() in MainWindow.
    """
    start_requested = pyqtSignal(dict)  # emitted with config dict when user clicks Start

    def __init__(self, parent=None):
        super().__init__(parent)
        self.scan_results = []  # [{freq_idx, R, X, Z, phase}]
        self.raw_rows = []
        self.scan_meta = {}

        # ── State for calibrated measurement (0Ω offset + DUT) ──
        self.calib_mode = False          # True while in calibrated two-phase mode
        self.calib_phase = None          # 'baseline' | 'dut' | None
        self.baseline_raw_rows = []      # 0Ω tooandmed (freq_idx -> list)
        self.baseline_meta = {}          # 0Ω metaandmed
        self.baseline_avg = {}           # {freq_idx: (avg_real, avg_imag)}
        self.output_folder = None        # output folder for CSV/Excel files
        self.pending_cfg = None          # config held between baseline and DUT phases
        self.baseline_csv_path = None
        self.dut_csv_path = None
        self.last_analysis_sheet = None
        self.calibration_path = None     # path to loaded calibration JSON
        self.calibration_data = None     # parsed calibration JSON content
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # ── Top: settings + progress ──
        top = QHBoxLayout()

        cfg_group = QGroupBox("Scan Settings")
        cfg_form = QFormLayout(cfg_group)
        cfg_form.setLabelAlignment(Qt.AlignRight)

        self.le_component = QLineEdit("")
        self.le_component.setPlaceholderText("e.g. 982.4")
        cfg_form.addRow("Reference resistor (Ω):", self.le_component)

        self.cb_current = QComboBox()
        for i in range(16):
            self.cb_current.addItem(f"[{i:2d}]  {format_current(i)}")
        self.cb_current.setCurrentIndex(12)
        cfg_form.addRow("Stim Current:", self.cb_current)

        self.cb_gain = QComboBox()
        for i, g in enumerate(GAINS):
            self.cb_gain.addItem(f"[{i}]  {g:.0f}x")
        cfg_form.addRow("Gain:", self.cb_gain)

        freq_row = QHBoxLayout()
        self.sp_freq_start = QSpinBox()
        self.sp_freq_start.setRange(0, 59)
        self.sp_freq_start.setValue(0)
        self.sp_freq_end = QSpinBox()
        self.sp_freq_end.setRange(0, 59)
        self.sp_freq_end.setValue(52)
        freq_row.addWidget(self.sp_freq_start)
        freq_row.addWidget(QLabel("to"))
        freq_row.addWidget(self.sp_freq_end)

        cfg_form.addRow("Frequencies F:", freq_row)

        # Calibrated mode controls (0Ω offset + DUT)
        self.cb_calib = QCheckBox("Calibrated mode (0Ω first, then DUT)")
        self.cb_calib.setChecked(True)
        self.cb_calib.setToolTip(
            "When enabled: first it asks for a 0Ω resistor and performs a baseline measurement.\n"
            "Then it asks for the device under test and performs BIS.\n"
            "Both measurements use EXACTLY THE SAME settings (current, gain, filters)."
        )
        cfg_form.addRow("", self.cb_calib)
        
        # Kalibratsiooni rakendamise sektsioon
        self.cb_apply_cal = QCheckBox("Apply calibration to analysis.xlsx")
        self.cb_apply_cal.setChecked(False)
        self.cb_apply_cal.setToolTip(
            "When enabled, a calibration JSON is loaded and the analysis.xlsx\n"
            "file gets additional tables (5-6) with calibrated results."
        )
        self.cb_apply_cal.toggled.connect(self._on_apply_cal_toggled)
        cfg_form.addRow("", self.cb_apply_cal)

        cal_row = QHBoxLayout()
        self.le_cal_path = QLineEdit()
        self.le_cal_path.setReadOnly(True)
        self.le_cal_path.setPlaceholderText("No calibration file selected")
        self.btn_cal_browse = QPushButton("...")
        self.btn_cal_browse.setMaximumWidth(40)
        self.btn_cal_browse.clicked.connect(self._browse_calibration)
        self.btn_cal_remove = QPushButton("✕")
        self.btn_cal_remove.setMaximumWidth(30)
        self.btn_cal_remove.setToolTip("Remove calibration file")
        self.btn_cal_remove.setStyleSheet(
            f"QPushButton {{ color: {COLORS['red']}; border-color: {COLORS['red']}; }}"
            f"QPushButton:hover {{ background: {COLORS['red']}; color: white; }}"
        )
        self.btn_cal_remove.clicked.connect(self._remove_calibration)
        cal_row.addWidget(self.le_cal_path, 1)
        cal_row.addWidget(self.btn_cal_browse)
        cal_row.addWidget(self.btn_cal_remove)
        cfg_form.addRow("Calibration JSON:", cal_row)
            
        top.addWidget(cfg_group)

        # ── Advanced Settings ──
        adv_group = QGroupBox("Advanced Settings")
        adv_form = QFormLayout(adv_group)
        adv_form.setLabelAlignment(Qt.AlignRight)
        adv_form.setSpacing(6)

        # Analoog HPF
        self.cb_ahpf = QComboBox()
        ahpf_opts = [
            "[0]  100 Hz", "[1]  200 Hz", "[2]  500 Hz", "[3]  1 kHz",
            "[4]  2 kHz", "[5]  5 kHz", "[6]  10 kHz",
            "[7]  Bypass (res. open)",
            "[8]  42.4 MΩ + cap short", "[9]  21.2 MΩ + cap short",
            "[10] 8.4 MΩ + cap short", "[11] 4.2 MΩ + cap short",
            "[12] 2.2 MΩ + cap short", "[13] 848 kΩ + cap short",
            "[14] 848 kΩ + cap short",
            "[15] Full bypass (res. open + cap short)",
        ]
        for o in ahpf_opts:
            self.cb_ahpf.addItem(o)
        self.cb_ahpf.setCurrentIndex(15)
        adv_form.addRow("Analog HPF:", self.cb_ahpf)

        # Digitaalfilter HPF
        self.cb_dhpf = QComboBox()
        for idx, lbl in enumerate(["[0]  Bypass", "[1]  0.00025 × SR", "[2]  0.002 × SR", "[3]  0.002 × SR"]):
            self.cb_dhpf.addItem(lbl)
        self.cb_dhpf.setCurrentIndex(0)
        adv_form.addRow("Digital HPF:", self.cb_dhpf)

        # Digitaalfilter LPF
        self.cb_dlpf = QComboBox()
        for idx, lbl in enumerate(["[0]  Bypass", "[1]  0.005 × SR", "[2]  0.02 × SR", "[3]  0.08 × SR",
                                    "[4]  0.25 × SR", "[5]  0.25 × SR", "[6]  0.25 × SR", "[7]  0.25 × SR"]):
            self.cb_dlpf.addItem(lbl)
        self.cb_dlpf.setCurrentIndex(0)
        adv_form.addRow("Digital LPF:", self.cb_dlpf)

        # DC Restore
        self.cb_dcrestore = QComboBox()
        self.cb_dcrestore.addItem("[0]  Off (default)")
        self.cb_dcrestore.addItem("[1]  On (10 MΩ feedback)")
        self.cb_dcrestore.setCurrentIndex(0)
        adv_form.addRow("DC Restore:", self.cb_dcrestore)

        # Ext Cap
        self.cb_extcap = QComboBox()
        self.cb_extcap.addItem("[0]  Off (default)")
        self.cb_extcap.addItem("[1]  On (CEXT DRVXC-DRVSJ)")
        self.cb_extcap.setCurrentIndex(0)
        adv_form.addRow("Ext Cap:", self.cb_extcap)

        # Amp Range (IDRV_RGE)
        self.cb_amprange = QComboBox()
        for idx, lbl in enumerate(["[0]  Low", "[1]  Medium-Low", "[2]  Medium-High", "[3]  High"]):
            self.cb_amprange.addItem(lbl)
        self.cb_amprange.setCurrentIndex(3)
        adv_form.addRow("Amp Range:", self.cb_amprange)

        # Amp BW
        self.cb_ampbw = QComboBox()
        for idx, lbl in enumerate(["[0]  Low", "[1]  Medium-Low", "[2]  Medium-High", "[3]  High"]):
            self.cb_ampbw.addItem(lbl)
        self.cb_ampbw.setCurrentIndex(3)
        adv_form.addRow("Amp BW:", self.cb_ampbw)

        layout.addWidget(adv_group)

        # Progress paneel
        prog_group = QGroupBox("Progress")
        prog_layout = QVBoxLayout(prog_group)

        self.lbl_status = QLabel("Idle")
        self.lbl_status.setStyleSheet(f"color: {COLORS['text_dim']}; font-size: 13px;")
        self.lbl_status.setAlignment(Qt.AlignCenter)

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.progress.setMinimumHeight(20)

        self.lbl_freq_current = QLabel("—")
        self.lbl_freq_current.setStyleSheet(f"color: {COLORS['accent2']}; font-size: 11px;")

        self.lbl_eta = QLabel("")
        self.lbl_eta.setStyleSheet(f"color: {COLORS['text_dim']}; font-size: 11px;")

        btn_row = QHBoxLayout()
        self.btn_start = QPushButton("▶  Start Scan")
        self.btn_start.setObjectName("btn_start")
        self.btn_start.setMinimumHeight(36)
        self.btn_start.clicked.connect(self._on_start)

        self.btn_stop = QPushButton("■  Stop")
        self.btn_stop.setObjectName("btn_stop")
        self.btn_stop.setMinimumHeight(36)
        self.btn_stop.setEnabled(False)

        btn_row.addWidget(self.btn_start)
        btn_row.addWidget(self.btn_stop)

        prog_layout.addWidget(self.lbl_status)
        prog_layout.addWidget(self.progress)
        prog_layout.addWidget(self.lbl_freq_current)
        prog_layout.addWidget(self.lbl_eta)
        prog_layout.addStretch()
        prog_layout.addLayout(btn_row)
        top.addWidget(prog_group)
        layout.addLayout(top)

        # ── Resultte tabel ──
        res_group = QGroupBox("BIS Results")
        res_layout = QVBoxLayout(res_group)

        self.tbl_results = QTableWidget(0, 10)
        self.tbl_results.setHorizontalHeaderLabels(
            ['F idx', 'Frequency', 'OSR', 'R (Ω)', 'X (Ω)', '|Z| (Ω)', 'Phase (°)',
             'R_cal (Ω)', '|Z|_cal (Ω)', 'Phase_cal (°)'])
        self.tbl_results.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl_results.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_results.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_results.setAlternatingRowColors(True)
        self.tbl_results.setStyleSheet(
            self.tbl_results.styleSheet() +
            f"QTableWidget::item:alternate {{ background: {COLORS['bg2']}; }}"
        )

        btn_row2 = QHBoxLayout()
        self.btn_export = QPushButton("📄  Export CSV")
        self.btn_export.clicked.connect(self._export_csv)
        self.btn_clear = QPushButton("Clear")
        self.btn_clear.clicked.connect(self._clear)

        btn_row2.addWidget(self.btn_export)
        btn_row2.addWidget(self.btn_clear)
        btn_row2.addStretch()

        res_layout.addWidget(self.tbl_results)
        res_layout.addLayout(btn_row2)
        layout.addWidget(res_group)

        self._start_time = None
    
    def _on_apply_cal_toggled(self, checked):
        """Toggle handler — enable/disable calibration."""
        if checked and not self.calibration_path:
            self._browse_calibration()
        if checked and not self.calibration_path:
            # Kui ei valitud, eemalda check
            self.cb_apply_cal.setChecked(False)

    def _browse_calibration(self):
        """Open file dialog to select a calibration JSON."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select calibration JSON",
            self.calibration_path or "",
            "Calibration JSON (*.json);;All files (*)"
        )
        if not path:
            return
        try:
            import json
            with open(path) as f:
                cal = json.load(f)
            # Lihtne sanity-check
            if 'calibration' not in cal or 'config' not in cal:
                raise ValueError("Vigane JSON struktuur")
            self.calibration_path = path
            self.calibration_data = cal
            # Show brief summary
            ref = cal.get('reference_resistor_ohm', '?')
            n = len(cal.get('calibration', {}))
            self.le_cal_path.setText(f"{os.path.basename(path)}  (ref={ref}Ω, n_freq={n})")
        except Exception as e:
            QMessageBox.critical(self, "Calibration load error",
                                 f"Could not load calibration:\n\n{e}")
            self.calibration_path = None
            self.calibration_data = None
            self.le_cal_path.clear()
            self.cb_apply_cal.setChecked(False)

    def _remove_calibration(self):
        """Remove the currently loaded calibration file."""
        self.calibration_path = None
        self.calibration_data = None
        self.le_cal_path.clear()
        self.cb_apply_cal.setChecked(False)

    def _check_calibration_compatibility(self, dut_meta):
        """Check that the current measurement config matches the loaded calibration.
        Tagasta (is_compatible: bool, warnings: list[str])."""
        if not self.calibration_data:
            return False, ['Calibration not loaded']
        warnings = []
        cfg = self.calibration_data['config']
        for key in ['stim_idx', 'gain_idx', 'ahpf_idx', 'dhpf_idx', 'dlpf_idx']:
            cal_v = cfg.get(key)
            dut_v = dut_meta.get(key)
            if isinstance(dut_v, str):
                try: dut_v = int(dut_v)
                except ValueError: pass
            if cal_v is not None and dut_v is not None and cal_v != dut_v:
                warnings.append(f"{key}: calibration={cal_v}, measurement={dut_v}")
        return (len(warnings) == 0), warnings
    
    def _on_start(self):
        if self.sp_freq_start.value() > self.sp_freq_end.value():
            return
        nom_text = self.le_component.text().strip().replace(',', '.')
        nominal = parse_nominal_ohm(nom_text)

        base_cfg = {
            'mode':         'scan',
            'stim_idx':     self.cb_current.currentIndex(),
            'gain_idx':     self.cb_gain.currentIndex(),
            'freq_start':   self.sp_freq_start.value(),
            'freq_end':     self.sp_freq_end.value(),
            'ahpf':         self.cb_ahpf.currentIndex(),
            'dhpf':         self.cb_dhpf.currentIndex(),
            'dlpf':         self.cb_dlpf.currentIndex(),
            'dcrestore':    self.cb_dcrestore.currentIndex(),
            'extcap':       self.cb_extcap.currentIndex(),
            'amprange':     self.cb_amprange.currentIndex(),
            'ampbw':        self.cb_ampbw.currentIndex(),
        }
        
        # Validate calibration before starting measurement
        if self.cb_apply_cal.isChecked():
            if not self.calibration_data:
                QMessageBox.warning(self, "Calibration not loaded",
                    "You enabled 'Apply calibration' but no JSON file is selected.")
                return
            # Check config match
            cfg = self.calibration_data['config']
            mismatches = []
            for key, my_val in [('stim_idx', base_cfg['stim_idx']),
                                ('gain_idx', base_cfg['gain_idx']),
                                ('ahpf_idx', base_cfg['ahpf']),
                                ('dhpf_idx', base_cfg['dhpf']),
                                ('dlpf_idx', base_cfg['dlpf'])]:
                cal_val = cfg.get(key)
                if cal_val is not None and cal_val != my_val:
                    mismatches.append(f"{key}: calibration={cal_val}, measurement={my_val}")
            if mismatches:
                msg = ("Calibration configuration does NOT match the current "
                       "measurement:\n\n" + "\n".join(mismatches) +
                       "\n\nContinue anyway?  (results may be inaccurate)")
                reply = QMessageBox.warning(self, "Config mismatch",
                                            msg, QMessageBox.Yes | QMessageBox.No)
                if reply != QMessageBox.Yes:
                    return
        
        if self.cb_calib.isChecked():
            # ── Calibrated mode: first ask for output folder, then 0Ω baseline, then DUT ──
            if not nom_text:
                QMessageBox.warning(self, "Missing information",
                    "Enter the nominal value of the resistor under test (e.g. 982.4 or 2p001K).")
                return

            # Check openpyxl availability early — if missing, Excel export will fail later.
            try:
                import openpyxl  # noqa: F401
            except ImportError:
                QMessageBox.critical(self, "openpyxl is missing",
                    "Calibrated mode requires the openpyxl package for Excel "
                    "analysis.\n\n"
                    "Open a terminal and run:\n"
                    "    pip install openpyxl\n\n"
                    "Then restart the GUI.")
                return

            # Ask for output folder
            folder = QFileDialog.getExistingDirectory(
                self, "Choose a folder for measurements of this configuration (or create a new one)",
                self.output_folder or "",
                QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks)
            if not folder:
                return
            self.output_folder = folder
            self.baseline_raw_rows = []
            self.baseline_meta = {}
            self.baseline_avg = {}
            self.baseline_csv_path = None
            self.dut_csv_path = None
            self.last_analysis_sheet = None

            # Store DUT info for later
            self.pending_cfg = dict(base_cfg)
            self.pending_cfg['component']    = nom_text
            self.pending_cfg['nominal_ohm']  = nominal if nominal is not None else nom_text

            # Confirmation dialog before 0Ω baseline scan
            reply = QMessageBox.question(
                self, "Step 1/2 — 0Ω baseline",
                "Connect the 0Ω resistor (or short) now.\n\n"
                f"Settings: I={format_current(base_cfg['stim_idx'])}  "
                f"G={GAINS[base_cfg['gain_idx']]:.0f}x  "
                f"F={base_cfg['freq_start']}..{base_cfg['freq_end']}\n\n"
                "Note: the 0Ω measurement uses EXACTLY the same settings as the "
                f"DUT ({nom_text} Ω) measurement — this is required so that "
                "the offset can be subtracted correctly.\n\n"
                "Continue?",
                QMessageBox.Ok | QMessageBox.Cancel)
            if reply != QMessageBox.Ok:
                self.pending_cfg = None
                return

            # Start baseline measurement
            self.calib_mode = True
            self.calib_phase = 'baseline'
            cfg = dict(base_cfg)
            cfg['component']   = '0'
            cfg['nominal_ohm'] = 0.0
            cfg['_calib_phase'] = 'baseline'
            self.start_requested.emit(cfg)
        else:
            # Regular single-pass scan (non-calibrated mode)
            self.calib_mode = False
            self.calib_phase = None
            cfg = dict(base_cfg)
            cfg['component']    = nom_text or 'unknown'
            cfg['nominal_ohm']  = nominal if nominal is not None else (nom_text or 'unknown')
            self.start_requested.emit(cfg)

    def start_dut_phase(self):
        """After 0Ω baseline completes: prompt for DUT label and start DUT scan."""
        if not self.pending_cfg:
            return
        nom_text = self.pending_cfg['component']
        reply = QMessageBox.question(
            self, "Step 2/2 — Device Under Test",
            f"Baseline (0Ω) saved.\n\n"
            f"Connect the device under test now: {nom_text} Ω\n\n"
            "The settings remain the same (all parameters are identical to the baseline).\n\n"
            "Continue?",
            QMessageBox.Ok | QMessageBox.Cancel)
        if reply != QMessageBox.Ok:
            self.calib_mode = False
            self.calib_phase = None
            self.pending_cfg = None
            self.set_running(False)
            return

        # Clear results table for DUT scan (baseline is kept in memory)
        self.tbl_results.setRowCount(0)
        self.scan_results = []
        self.raw_rows = []

        self.calib_phase = 'dut'
        cfg = dict(self.pending_cfg)
        cfg['_calib_phase'] = 'dut'
        self.start_requested.emit(cfg)

    def set_running(self, running):
        self.btn_start.setEnabled(not running)
        self.btn_stop.setEnabled(running)
        if running:
            self.lbl_status.setText("⏳ Scan running...")
            self.lbl_status.setStyleSheet(f"color: {COLORS['yellow']}; font-size: 13px;")
            self._start_time = time.time()
        else:
            self.lbl_status.setText("✓ Ready")
            self.lbl_status.setStyleSheet(f"color: {COLORS['green']}; font-size: 13px;")

    def update_progress(self, done, total):
        pct = int(done * 100 / total) if total > 0 else 0
        self.progress.setValue(pct)
        self.progress.setFormat(f"{done}/{total} frequencies  ({pct}%)")
        if self._start_time and done > 0:
            elapsed = time.time() - self._start_time
            eta = elapsed / done * (total - done)
            self.lbl_eta.setText(f"ETA: ~{eta:.0f}s")

    def add_freq_result(self, freq_idx, n_samples, R, X):
        import cmath as _cmath
        mag = math.sqrt(R*R + X*X)
        phase = math.degrees(math.atan2(X, R)) if mag > 0 else 0
        osr = get_adc_osr(freq_idx)
        freq_str = format_freq(FREQ_HZ[freq_idx]) if freq_idx < len(FREQ_HZ) else '?'

        # ── Calibrated values (if calibration JSON is loaded) ──
        R_cal = X_cal = mag_cal = phase_cal = None
        if (self.cb_apply_cal.isChecked() and self.calibration_data
                and self.calib_phase == 'dut' and self.baseline_avg):
            cal_k_map = {int(fi): complex(v['k_real'], v['k_imag'])
                         for fi, v in self.calibration_data['calibration'].items()}
            K = self.calibration_data['config']['K_value']
            k = cal_k_map.get(freq_idx)
            base_avg = self.baseline_avg.get(freq_idx)
            if k is not None and base_avg is not None:
                stim_idx = self.scan_meta.get('stim_idx', 12)
                gain_idx = self.scan_meta.get('gain_idx', 0)
                I_A = CURRENTS_UA[stim_idx] * 1e-6
                G = GAINS[gain_idx]
                # DUT ADC keskmised = R/X * K (inverse of component_from_adc)
                denom = IMP_DENOM_CONST * G * I_A
                dut_r_adc = R * denom
                dut_i_adc = X * denom
                base_r_adc, base_i_adc = base_avg[0], base_avg[1]
                net_r = dut_r_adc - base_r_adc
                net_i = dut_i_adc - base_i_adc
                z_uncal = complex(net_r, net_i) / K
                z_cal_c = k * z_uncal
                R_cal = z_cal_c.real
                X_cal = z_cal_c.imag
                mag_cal = abs(z_cal_c)
                phase_cal = math.degrees(_cmath.phase(z_cal_c)) if mag_cal > 0 else 0

        lbl = f"Last: F={freq_idx}  {freq_str}  →  |Z|={fmt3(mag)}Ω"
        if mag_cal is not None:
            lbl += f"  |Z|_cal={fmt3(mag_cal)}Ω"
        self.lbl_freq_current.setText(lbl)

        # Lisa tabelisse
        row = self.tbl_results.rowCount()
        self.tbl_results.insertRow(row)
        vals = [str(freq_idx), freq_str, str(osr),
                fmt3(R), fmt3(X), fmt3(mag), fmt3(phase),
                fmt3(R_cal) if R_cal is not None else '—',
                fmt3(mag_cal) if mag_cal is not None else '—',
                fmt3(phase_cal) if phase_cal is not None else '—']
        for col, val in enumerate(vals):
            item = QTableWidgetItem(val)
            item.setTextAlignment(Qt.AlignCenter)
            # Color coding by |Z| (column 5)
            if col == 5:
                if mag < 500:
                    item.setForeground(QColor(COLORS['green']))
                elif mag < 5000:
                    item.setForeground(QColor(COLORS['yellow']))
                else:
                    item.setForeground(QColor(COLORS['red']))
            # Color coding by |Z|_cal (column 8)
            if col == 8 and mag_cal is not None:
                item.setForeground(QColor(COLORS['cyan']))
                item.setFont(QFont('Consolas', 10, QFont.Bold))
            self.tbl_results.setItem(row, col, item)
        self.tbl_results.scrollToBottom()

        stim_idx = self.scan_meta.get('stim_idx', 12)
        gain_idx = self.scan_meta.get('gain_idx', 0)
        K = get_K(freq_idx, gain_idx, stim_idx)
        self.scan_results.append({
            'freq_idx':     freq_idx,
            'freq_hz':      FREQ_HZ[freq_idx] if freq_idx < len(FREQ_HZ) else 0,
            'osr':          get_adc_osr(freq_idx),
            'K':            K,
            'stim_idx':     self.scan_meta.get('stim_idx', '?'),
            'gain_idx':     self.scan_meta.get('gain_idx', '?'),
            'I_uA':         CURRENTS_UA[stim_idx],
            'G':            GAINS[gain_idx],
            'real_adc_avg': fmt3(R * K),
            'imag_adc_avg': fmt3(X * K),
            'R_ohm':        R,
            'X_ohm':        X,
            'Z_ohm':        mag,
            'phase_deg':    phase,
            'n_samples':    n_samples,
            'oor_count':    0,
            'R_cal':        R_cal if R_cal is not None else '',
            'X_cal':        X_cal if X_cal is not None else '',
            'Z_cal':        mag_cal if mag_cal is not None else '',
            'phase_cal':    phase_cal if phase_cal is not None else '',
        })

    def _export_csv(self):
        if not self.raw_rows and not self.scan_results:
            QMessageBox.information(self, "Info", "Resulted puuduvad.")
            return
        fname, _ = QFileDialog.getSaveFileName(
            self, "Save Raw ADC CSV",
            (lambda nom: f"{nom}ohm_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_raw.csv")(
                str(self.scan_meta.get('nominal_ohm') or self.scan_meta.get('component','scan')).replace('.','p')),
            "CSV Files (*.csv)")
        if not fname:
            return

        meta = self.scan_meta
        nominal = meta.get('nominal_ohm')

        # ── Raw ADC file ──
        lines_raw = [
            "# MAX30009 Raw ADC data",
            f"# timestamp: {datetime.datetime.now().isoformat()}",
            f"# nominal_ohm: {meta.get('nominal_ohm','?')}",
        ]
        lines_raw += [
            f"# stim_idx: {meta.get('stim_idx','?')}",
            f"# stim_uA: {CURRENTS_UA[meta.get('stim_idx',12)]}",
            f"# gain_idx: {meta.get('gain_idx','?')}",
            f"# gain_x: {GAINS[meta.get('gain_idx',0)]:.0f}",
            f"# freq_start: {meta.get('freq_start','?')}",
            f"# freq_end: {meta.get('freq_end','?')}",
            f"# ahpf_idx: {meta.get('ahpf',15)}",
            f"# dhpf_idx: {meta.get('dhpf',0)}",
            f"# dlpf_idx: {meta.get('dlpf',0)}",
            f"# dcrestore: {meta.get('dcrestore',0)}",
            f"# extcap: {meta.get('extcap',0)}",
            f"# amprange: {meta.get('amprange',3)}",
            f"# ampbw: {meta.get('ampbw',3)}",
            f"# K_osr128: {K_BY_ADC_OSR[128]}",
            f"# K_osr256: {K_BY_ADC_OSR[256]}",
            f"# K_osr512: {K_BY_ADC_OSR[512]}",
            f"# K_osr1024: {K_BY_ADC_OSR[1024]}",
            "# columns: freq_idx, freq_hz, adc_osr, sample_nr, real_adc, imag_adc, flags",
            "# flags: bit0=idrv_ovr bit1=code_ovr bit2=code_und bit3=leadsoff bit4=error bit5=marker",
            "#",
        ]
        with open(fname, 'w', newline='', encoding='utf-8') as f:
            for ln in lines_raw:
                f.write(ln + "\n")
            writer = csv.DictWriter(f, fieldnames=[
                'freq_idx', 'freq_hz', 'adc_osr', 'sample_nr', 'real_adc', 'imag_adc', 'flags'])
            writer.writeheader()
            for r in self.raw_rows:
                r2 = dict(r)
                r2['adc_osr'] = get_adc_osr(r['freq_idx'])
                writer.writerow(r2)

        # ── Averaged impedance file ──
        avg_fname = fname.replace('_raw.csv', '_impedance.csv')
        imp_header = [
            "# MAX30009 Impedance results (averaged)",
            f"# timestamp: {datetime.datetime.now().isoformat()}",
            f"# nominal_ohm: {meta.get('nominal_ohm','?')}",
        ]
        imp_header += [
            f"# stim_idx: {meta.get('stim_idx','?')}",
            f"# stim_uA: {CURRENTS_UA[meta.get('stim_idx',12)]}",
            f"# gain_idx: {meta.get('gain_idx','?')}",
            f"# gain_x: {GAINS[meta.get('gain_idx',0)]:.0f}",
            f"# freq_start: {meta.get('freq_start','?')}",
            f"# freq_end: {meta.get('freq_end','?')}",
            f"# ahpf_idx: {meta.get('ahpf',15)}",
            f"# dhpf_idx: {meta.get('dhpf',0)}",
            f"# dlpf_idx: {meta.get('dlpf',0)}",
            f"# dcrestore: {meta.get('dcrestore',0)}",
            f"# extcap: {meta.get('extcap',0)}",
            f"# amprange: {meta.get('amprange',3)}",
            f"# ampbw: {meta.get('ampbw',3)}",
            "#",
        ]
        with open(avg_fname, 'w', newline='', encoding='utf-8') as f:
            for ln in imp_header:
                f.write(ln + "\n")
            writer = csv.DictWriter(f, extrasaction='ignore', fieldnames=[
                'freq_idx', 'freq_hz', 'osr', 'K', 'stim_idx', 'gain_idx', 'I_uA', 'G',
                'real_adc_avg', 'imag_adc_avg', 'R_ohm', 'X_ohm', 'Z_ohm', 'phase_deg',
                'R_cal', 'X_cal', 'Z_cal', 'phase_cal',
                'n_samples', 'oor_count'])
            writer.writeheader()
            writer.writerows(self.scan_results)

        n_raw = len(self.raw_rows)
        n_avg = len(self.scan_results)
        QMessageBox.information(self, "Saved",
            f"Raw ADC ({n_raw} rows):\n{fname}\n\n"
            f"Impedance ({n_avg} frequencies):\n{avg_fname}")

    def _clear(self):
        self.tbl_results.setRowCount(0)
        self.scan_results.clear()
        self.raw_rows.clear()
        self.progress.setValue(0)
        self.lbl_freq_current.setText("—")
        self.lbl_eta.setText("")
        self.lbl_status.setText("Idle")
        self.lbl_status.setStyleSheet(f"color: {COLORS['text_dim']}; font-size: 13px;")

    def _auto_export(self, log_widget=None):
        """Automatic CSV export after scan — asks for output folder."""
        if not self.raw_rows and not self.scan_results:
            return

        meta = self.scan_meta
        nom = str(meta.get('nominal_ohm') or meta.get('component', 'scan')).replace('.', 'p')
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

        folder = QFileDialog.getExistingDirectory(
            self, "Select output folder", "",
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks)
        if not folder:
            if log_widget:
                log_widget.append("  ℹ CSV export cancelled.")
            return

        fname = os.path.join(folder, f"{nom}ohm_{ts}_raw.csv")
        avg_fname = os.path.join(folder, f"{nom}ohm_{ts}_impedance.csv")

        try:
            lines_raw = [
                "# MAX30009 Raw ADC data",
                f"# timestamp: {datetime.datetime.now().isoformat()}",
                f"# nominal_ohm: {meta.get('nominal_ohm','?')}",
                f"# stim_idx: {meta.get('stim_idx','?')}",
                f"# stim_uA: {CURRENTS_UA[meta.get('stim_idx',12)]}",
                f"# gain_idx: {meta.get('gain_idx','?')}",
                f"# gain_x: {GAINS[meta.get('gain_idx',0)]:.0f}",
                f"# freq_start: {meta.get('freq_start','?')}",
                f"# freq_end: {meta.get('freq_end','?')}",
                f"# ahpf_idx: {meta.get('ahpf',15)}",
                f"# dhpf_idx: {meta.get('dhpf',0)}",
                f"# dlpf_idx: {meta.get('dlpf',0)}",
                f"# dcrestore: {meta.get('dcrestore',0)}",
                f"# extcap: {meta.get('extcap',0)}",
                f"# amprange: {meta.get('amprange',3)}",
                f"# ampbw: {meta.get('ampbw',3)}",
                f"# K_osr128: {K_BY_ADC_OSR[128]}",
                f"# K_osr256: {K_BY_ADC_OSR[256]}",
                f"# K_osr512: {K_BY_ADC_OSR[512]}",
                f"# K_osr1024: {K_BY_ADC_OSR[1024]}",
                "# columns: freq_idx, freq_hz, adc_osr, sample_nr, real_adc, imag_adc, flags",
                "# flags: bit0=idrv_ovr bit1=code_ovr bit2=code_und bit3=leadsoff bit4=error bit5=marker",
                "#",
            ]
            with open(fname, 'w', newline='', encoding='utf-8') as f:
                for ln in lines_raw:
                    f.write(ln + "\n")
                writer = csv.DictWriter(f, fieldnames=[
                    'freq_idx', 'freq_hz', 'adc_osr', 'sample_nr', 'real_adc', 'imag_adc', 'flags'])
                writer.writeheader()
                for r in self.raw_rows:
                    r2 = dict(r)
                    r2['adc_osr'] = get_adc_osr(r['freq_idx'])
                    writer.writerow(r2)

            imp_header = [
                "# MAX30009 Impedance results (averaged)",
                f"# timestamp: {datetime.datetime.now().isoformat()}",
                f"# nominal_ohm: {meta.get('nominal_ohm','?')}",
                f"# stim_idx: {meta.get('stim_idx','?')}",
                f"# stim_uA: {CURRENTS_UA[meta.get('stim_idx',12)]}",
                f"# gain_idx: {meta.get('gain_idx','?')}",
                f"# gain_x: {GAINS[meta.get('gain_idx',0)]:.0f}",
                f"# freq_start: {meta.get('freq_start','?')}",
                f"# freq_end: {meta.get('freq_end','?')}",
                f"# ahpf_idx: {meta.get('ahpf',15)}",
                f"# K_osr128: {K_BY_ADC_OSR[128]}",
                f"# K_osr256: {K_BY_ADC_OSR[256]}",
                f"# K_osr512: {K_BY_ADC_OSR[512]}",
                f"# K_osr1024: {K_BY_ADC_OSR[1024]}",
                "#",
            ]
            with open(avg_fname, 'w', newline='', encoding='utf-8') as f:
                for ln in imp_header:
                    f.write(ln + "\n")
                writer = csv.DictWriter(f, extrasaction='ignore', fieldnames=[
                    'freq_idx', 'freq_hz', 'osr', 'K', 'stim_idx', 'gain_idx', 'I_uA', 'G',
                    'real_adc_avg', 'imag_adc_avg', 'R_ohm', 'X_ohm', 'Z_ohm', 'phase_deg',
                    'R_cal', 'X_cal', 'Z_cal', 'phase_cal',
                    'n_samples', 'oor_count'])
                writer.writeheader()
                writer.writerows(self.scan_results)

            msg = (f"  ✓ Raw ({len(self.raw_rows)} rows): {fname}\n"
                   f"  ✓ Impedance ({len(self.scan_results)} freqs): {avg_fname}")
            if log_widget:
                log_widget.append(msg)
        except Exception as e:
            err = f"  ✗ CSV export error: {e}"
            if log_widget:
                log_widget.append(err)

    # ──────────────────────────────────────────────────────────
    # Calibrated mode helper functions
    # ──────────────────────────────────────────────────────────
    def _compute_avg_from_raw(self, raw_rows):
        """raw_rows → {freq_idx: (avg_real, avg_imag, n_samples, freq_hz)}"""
        by_f = {}
        for r in raw_rows:
            fi = r['freq_idx']
            if fi not in by_f:
                by_f[fi] = {'real': [], 'imag': [], 'hz': r.get('freq_hz', 0)}
            by_f[fi]['real'].append(r['real_adc'])
            by_f[fi]['imag'].append(r['imag_adc'])
        out = {}
        for fi, d in by_f.items():
            if d['real']:
                out[fi] = (
                    sum(d['real']) / len(d['real']),
                    sum(d['imag']) / len(d['imag']),
                    len(d['real']),
                    d['hz'],
                )
        return out


    def _save_raw_csv(self, raw_rows, meta, folder, label=None, log_widget=None):
        """Save one measurement's raw ADC rows to a separate CSV file."""
        if not folder or not raw_rows:
            return None

        nominal = meta.get('nominal_ohm')
        if label is None:
            if nominal is None:
                label = str(meta.get('component', 'scan'))
            else:
                label = str(nominal)

        safe_label = str(label).strip().replace(',', '.').replace(' ', '_')
        safe_label = safe_label.replace('.', 'p')
        safe_label = re.sub(r'[^0-9A-Za-z_\-]+', '_', safe_label)
        if not safe_label:
            safe_label = 'scan'

        if 'ohm' not in safe_label.lower():
            safe_label = f"{safe_label}ohm"

        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        raw_path = os.path.join(folder, f"{safe_label}_{ts}_raw.csv")

        header_lines = [
            "# MAX30009 Raw ADC data",
            f"# timestamp: {datetime.datetime.now().isoformat()}",
            f"# nominal_ohm: {meta.get('nominal_ohm')}",
            f"# component: {meta.get('component')}",
            f"# calib_phase: {meta.get('calib_phase')}",
            f"# stim_idx: {meta.get('stim_idx')}",
            f"# stim_uA: {CURRENTS_UA[meta.get('stim_idx', 12)]}",
            f"# gain_idx: {meta.get('gain_idx')}",
            f"# gain_x: {GAINS[meta.get('gain_idx', 0)]}",
            f"# freq_start: {meta.get('freq_start')}",
            f"# freq_end: {meta.get('freq_end')}",
            f"# ahpf_idx: {meta.get('ahpf', 15)}",
            f"# dhpf_idx: {meta.get('dhpf', 0)}",
            f"# dlpf_idx: {meta.get('dlpf', 0)}",
            f"# dcrestore: {meta.get('dcrestore', 0)}",
            f"# extcap: {meta.get('extcap', 0)}",
            f"# amprange: {meta.get('amprange', 3)}",
            f"# ampbw: {meta.get('ampbw', 3)}",
            f"# K_osr128: {K_BY_ADC_OSR[128]}",
            f"# K_osr256: {K_BY_ADC_OSR[256]}",
            f"# K_osr512: {K_BY_ADC_OSR[512]}",
            f"# K_osr1024: {K_BY_ADC_OSR[1024]}",
            "# columns: freq_idx, freq_hz, adc_osr, sample_nr, real_adc, imag_adc, flags",
            "# flags: bit0=idrv_ovr bit1=code_ovr bit2=code_und bit3=leadsoff bit4=error bit5=marker",
            "#",
        ]

        with open(raw_path, 'w', newline='', encoding='utf-8') as f:
            for line in header_lines:
                f.write(line + "\n")
            writer = csv.writer(f)
            writer.writerow(['freq_idx', 'freq_hz', 'adc_osr', 'sample_nr', 'real_adc', 'imag_adc', 'flags'])
            for row in raw_rows:
                fi = row.get('freq_idx')
                writer.writerow([
                    fi,
                    row.get('freq_hz'),
                    get_adc_osr(fi) if fi is not None else '',
                    row.get('sample_nr'),
                    row.get('real_adc'),
                    row.get('imag_adc'),
                    row.get('flags'),
                ])

        if log_widget:
            log_widget.append(f"  ✓ Raw CSV saved: {raw_path}")
        return raw_path

    def _build_measurement_rows(self, raw_rows, meta):
        avg_map = self._compute_avg_from_raw(raw_rows)
        stim_idx = meta.get('stim_idx', 12)
        gain_idx = meta.get('gain_idx', 0)
        I_A = CURRENTS_UA[stim_idx] * 1e-6
        G = GAINS[gain_idx]

        rows = []
        for fi in sorted(avg_map.keys()):
            avg_r, avg_i, n_samples, freq_hz = avg_map[fi]
            mag_adc = math.sqrt(avg_r * avg_r + avg_i * avg_i)
            z_ohm = mag_adc / (IMP_DENOM_CONST * G * I_A) if I_A and G else math.nan
            phase_deg = math.degrees(math.atan2(avg_i, avg_r)) if mag_adc > 0 else 0.0
            rows.append({
                'freq_idx': fi,
                'freq_hz': freq_hz,
                'osr': get_adc_osr(fi),
                'real_adc_avg': avg_r,
                'imag_adc_avg': avg_i,
                'mag_adc': mag_adc,
                'Z_ohm': z_ohm,
                'phase_deg': phase_deg,
                'n_samples': n_samples,
            })
        return rows

    def _ensure_config_sheet(self, wb, meta):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if 'Configuration' in wb.sheetnames:
            del wb['Configuration']
        ws = wb.create_sheet('Configuration', 0)

        title_fill = PatternFill('solid', fgColor='1F4E78')
        section_fill = PatternFill('solid', fgColor='D9EAF7')
        border = Border(
            left=Side(style='thin', color='C9D1D9'),
            right=Side(style='thin', color='C9D1D9'),
            top=Side(style='thin', color='C9D1D9'),
            bottom=Side(style='thin', color='C9D1D9'),
        )

        ws.merge_cells('A1:D1')
        c = ws['A1']
        c.value = 'MAX30009 BIS – Configuration'
        c.font = Font(bold=True, size=14, color='FFFFFF')
        c.fill = title_fill
        c.alignment = Alignment(horizontal='center')

        info_rows = [
            ('Stim idx', meta.get('stim_idx')),
            ('Stim current (µA)', CURRENTS_UA[meta.get('stim_idx', 12)]),
            ('Gain idx', meta.get('gain_idx')),
            ('Gain (x)', GAINS[meta.get('gain_idx', 0)]),
            ('Freq idx beginning', meta.get('freq_start')),
            ('Freq idx end', meta.get('freq_end')),
            ('Freq beginning (Hz)', FREQ_HZ[meta.get('freq_start')] if isinstance(meta.get('freq_start'), int) and 0 <= meta.get('freq_start') < len(FREQ_HZ) else ''),
            ('Freq end (Hz)', FREQ_HZ[meta.get('freq_end')] if isinstance(meta.get('freq_end'), int) and 0 <= meta.get('freq_end') < len(FREQ_HZ) else ''),
            ('AHPF idx', meta.get('ahpf', 15)),
            ('DHPF idx', meta.get('dhpf', 0)),
            ('DLPF idx', meta.get('dlpf', 0)),
            ('DC Restore', meta.get('dcrestore', 0)),
            ('Ext Cap', meta.get('extcap', 0)),
            ('AmpRange', meta.get('amprange', 3)),
            ('AmpBW', meta.get('ampbw', 3)),
            ('Formula |Z|', '|Z| = MAG / (524288 · G · (2/π) · I)'),
            ('Note', 'All following sheets contain results for 0Ω + DUT measurement pairs made with the same configuration.'),
            ('Last updated', datetime.datetime.now().isoformat()),
            ('Measurement folder', self.output_folder or ''),
        ]

        ws['A3'] = 'Parameter'
        ws['B3'] = 'Value'
        ws['A3'].font = ws['B3'].font = Font(bold=True, color='FFFFFF')
        ws['A3'].fill = ws['B3'].fill = title_fill
        ws['A3'].alignment = ws['B3'].alignment = Alignment(horizontal='center')

        row = 4
        for key, value in info_rows:
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = section_fill
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', wrap_text=True)
            row += 1

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 4
        ws.column_dimensions['D'].width = 4
        ws.freeze_panes = 'A4'
        return ws

    def _make_sheet_name(self, nom):
        """Loob Exceli-sobiva baassheeti nime nominaaltakistusest."""
        if isinstance(nom, (int, float)):
            if nom >= 1e6:
                name = f"{nom / 1e6:.3f}MOhm"
            elif nom >= 1e3:
                name = f"{nom / 1e3:.3f}kOhm"
            elif nom == 0:
                name = '0Ohm'
            else:
                name = f"{nom:.3f}Ohm"
        else:
            name = str(nom or 'Measurement') + 'Ohm'
        for ch in ['\\', '/', '?', '*', '[', ']', ':']:
            name = name.replace(ch, '_')
        return name[:31]

    def _make_unique_sheet_name(self, wb, nom, timestamp_iso=None):
        base = self._make_sheet_name(nom)
        try:
            dt = datetime.datetime.fromisoformat(str(timestamp_iso)) if timestamp_iso else datetime.datetime.now()
        except Exception:
            dt = datetime.datetime.now()
        suffix = dt.strftime('%m%d_%H%M%S')
        max_base_len = max(1, 31 - len(suffix) - 1)
        name = f"{base[:max_base_len]}_{suffix}"
        counter = 2
        while name in wb.sheetnames:
            extra = f"_{counter}"
            trim_len = max(1, 31 - len(suffix) - len(extra) - 1)
            name = f"{base[:trim_len]}_{suffix}{extra}"
            counter += 1
        return name

    def _write_data_table(self, ws, start_row, title, headers, rows, *,
                          title_fill, header_fill, border, start_col=1,
                          number_formats=None):
        """number_formats: optional list same length as headers; each entry is
        an Excel number format string (e.g. '0.###', '0.###"%"') or None to
        leave default. Controls display precision without changing stored value."""
        from openpyxl.styles import Font, Alignment

        last_col = start_col + len(headers) - 1
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=last_col)
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, color='FFFFFF')
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='left')

        header_row = start_row + 1
        data_row = header_row + 1

        for idx, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=header_row, column=idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        for row_idx, values in enumerate(rows, start=data_row):
            for col_idx, value in enumerate(values, start=start_col):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                    if number_formats:
                        nf = number_formats[col_idx - start_col]
                        if nf:
                            cell.number_format = nf
                else:
                    cell.alignment = Alignment(horizontal='center')

        return data_row + len(rows)
    
    def _update_analysis_xlsx(self, baseline_raw, baseline_meta, dut_raw, dut_meta, log_widget=None):
        """Append a new measurement pair sheet to analysis.xlsx in the output folder.

        Sheet layout:
          Table 1 — 0Ω baseline: per-frequency averaged ADC + raw |Z|
          Table 2 — DUT: same columns
          Table 3 — Final result: DUT − baseline offset, |Z|_final, Δ vs nominal (Ω and %)

        Each call to this function adds one new sheet (named by nominal value + timestamp).
        The config sheet (sheet 0) is kept or created once.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

        folder = self.output_folder
        if not folder:
            raise RuntimeError('Output folder has not been selected.')

        xlsx_path = os.path.join(folder, 'analysis.xlsx')

        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        self._ensure_config_sheet(wb, dut_meta)

        baseline_rows = self._build_measurement_rows(baseline_raw, baseline_meta)
        dut_rows = self._build_measurement_rows(dut_raw, dut_meta)

        base_by_f = {row['freq_idx']: row for row in baseline_rows}
        dut_by_f = {row['freq_idx']: row for row in dut_rows}

        sheet_name = self._make_unique_sheet_name(wb, dut_meta.get('nominal_ohm'), dut_meta.get('timestamp'))
        ws = wb.create_sheet(sheet_name)

        title_fill = PatternFill('solid', fgColor='1F4E78')
        table_fill = PatternFill('solid', fgColor='2F75B5')
        header_fill = PatternFill('solid', fgColor='5B9BD5')
        border = Border(
            left=Side(style='thin', color='D0D7DE'),
            right=Side(style='thin', color='D0D7DE'),
            top=Side(style='thin', color='D0D7DE'),
            bottom=Side(style='thin', color='D0D7DE'),
        )

        ws.merge_cells('A1:I1')
        title = ws['A1']
        title.value = 'MAX30009 BIS – measurement pair'
        title.font = Font(bold=True, size=14, color='FFFFFF')
        title.fill = title_fill
        title.alignment = Alignment(horizontal='center')

        nom = dut_meta.get('nominal_ohm')
        nom_text = f"{nom} Ω" if nom is not None else str(dut_meta.get('component', 'DUT'))
        ws['A2'] = f"Baseline: 0 Ω    |    DUT: {nom_text}"
        ws['A3'] = f"Baseline time: {baseline_meta.get('timestamp', '-')}    |    DUT time: {dut_meta.get('timestamp', '-')}"

        row = 5

        common_headers = [
            'F idx', 'Frequency (Hz)', 'OSR',
            'Real ADC avg', 'Imag ADC avg', 'MAG ADC',
            '|Z| (Ω)', 'Phase (°)', 'n'
        ]

        # Number formats: max 3 decimal places displayed (trailing zeros trimmed),
        # but full precision is still stored in the cell.
        fmt_num  = '0.###'
        fmt_pct  = '0.###"%"'
        fmt_int  = '0'
        common_nfmt = [fmt_int, fmt_int, fmt_int,
                       fmt_num, fmt_num, fmt_num,
                       fmt_num, fmt_num, fmt_int]

        baseline_table_rows = []
        for item in baseline_rows:
            baseline_table_rows.append([
                item['freq_idx'],
                item['freq_hz'],
                item['osr'],
                item['real_adc_avg'],
                item['imag_adc_avg'],
                item['mag_adc'],
                item['Z_ohm'],
                item['phase_deg'],
                item['n_samples'],
            ])
        row = self._write_data_table(
            ws, row, 'Table 1 — 0Ω baseline all results', common_headers, baseline_table_rows,
            title_fill=table_fill, header_fill=header_fill, border=border,
            number_formats=common_nfmt
        )
        row += 2

        dut_table_rows = []
        for item in dut_rows:
            dut_table_rows.append([
                item['freq_idx'],
                item['freq_hz'],
                item['osr'],
                item['real_adc_avg'],
                item['imag_adc_avg'],
                item['mag_adc'],
                item['Z_ohm'],
                item['phase_deg'],
                item['n_samples'],
            ])
        row = self._write_data_table(
            ws, row, 'Table 2 — DUT all results', common_headers, dut_table_rows,
            title_fill=table_fill, header_fill=header_fill, border=border,
            number_formats=common_nfmt
        )
        row += 2

        conclusion_headers = [
            'F idx', 'Frequency (Hz)',
            '0Ω |Z| (Ω)', 'DUT |Z| (Ω)', 'Final |Z| (Ω)',
            'Nominaal (Ω)', 'Δ nom abs (Ω)', 'Δ nom (%)'
        ]
        conclusion_nfmt = [fmt_int, fmt_int,
                           fmt_num, fmt_num, fmt_num,
                           fmt_num, fmt_num, fmt_pct]

        conclusion_rows = []
        all_freqs = sorted(set(base_by_f.keys()) | set(dut_by_f.keys()))
        nominal_value = dut_meta.get('nominal_ohm')
        for fi in all_freqs:
            base = base_by_f.get(fi)
            dut = dut_by_f.get(fi)
            if not base or not dut:
                continue

            final_real = dut['real_adc_avg'] - base['real_adc_avg']
            final_imag = dut['imag_adc_avg'] - base['imag_adc_avg']
            final_mag = math.sqrt(final_real * final_real + final_imag * final_imag)

            stim_idx = dut_meta.get('stim_idx', 12)
            gain_idx = dut_meta.get('gain_idx', 0)
            I_A = CURRENTS_UA[stim_idx] * 1e-6
            G = GAINS[gain_idx]

            final_z = final_mag / (IMP_DENOM_CONST * G * I_A) if I_A and G else math.nan

            delta_nom_abs = ''
            delta_nom_pct = ''
            if isinstance(nominal_value, (int, float)):
                delta_nom_abs = abs(final_z - nominal_value)
                delta_nom_pct = (delta_nom_abs / abs(nominal_value) * 100.0) if nominal_value else 0.0

            conclusion_rows.append([
                fi,
                dut['freq_hz'],
                base['Z_ohm'],
                dut['Z_ohm'],
                final_z,
                nominal_value,
                delta_nom_abs if isinstance(delta_nom_abs, (int, float)) else '',
                delta_nom_pct if isinstance(delta_nom_pct, (int, float)) else '',
            ])

        conclusion_start = row
        row = self._write_data_table(
            ws, row,
            'Table 3 — Simple summary: measured impedances, final result, and comparison to nominal',
            conclusion_headers, conclusion_rows,
            title_fill=table_fill, header_fill=header_fill, border=border,
            number_formats=conclusion_nfmt
        )

        final_header = ws.cell(row=conclusion_start + 1, column=5)
        final_header.fill = PatternFill('solid', fgColor='70AD47')
        final_header.font = Font(bold=True, color='FFFFFF')
        final_col_fill = PatternFill('solid', fgColor='E2F0D9')
        final_col_font = Font(bold=True, color='006100')
        for data_row in range(conclusion_start + 2, row):
            cell = ws.cell(row=data_row, column=5)
            cell.fill = final_col_fill
            cell.font = final_col_font

        # ── Tables 5 and 6: apply calibration correction (if loaded) ──
        if self.cb_apply_cal.isChecked() and self.calibration_data:
            row += 2
            row = self._write_calibration_tables(
                ws, row, baseline_rows, dut_rows, dut_meta,
                title_fill=table_fill, header_fill=header_fill, border=border,
                fmt_num=fmt_num, fmt_pct=fmt_pct, fmt_int=fmt_int,
                log_widget=log_widget
            )

        widths = {
            'A': 8, 'B': 14, 'C': 12, 'D': 12, 'E': 13, 'F': 12, 'G': 14, 'H': 11, 'I': 14
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = 'A6'
        wb.save(xlsx_path)

        if log_widget:
            log_widget.append(f"  ✓ Excel uuendatud: {xlsx_path} (sheet: {sheet_name})")
    
        self.last_analysis_sheet = sheet_name
        return xlsx_path, sheet_name

    def _write_calibration_tables(self, ws, row, baseline_rows, dut_rows, dut_meta,
                               title_fill, header_fill, border,
                               fmt_num, fmt_pct, fmt_int, log_widget=None):
        """Write Table 5 (calibrated results per frequency) and Table 6
        (OSR-range statistics before vs after calibration).
        Returns the next free row number."""
        from openpyxl.styles import Font, PatternFill, Alignment
        import cmath

        cal = self.calibration_data
        cal_k = {int(fi): complex(v['k_real'], v['k_imag'])
                 for fi, v in cal['calibration'].items()}
        K = cal['config']['K_value']

        # Validate configuration
        is_compat, warnings = self._check_calibration_compatibility(dut_meta)
        if warnings and log_widget:
            for w in warnings:
                log_widget.append(f"  ⚠ Calibration config mismatch: {w}")

        # Kui konfig erineb, lisa sheet-i hoiatus
        if warnings:
            warn_cell = ws.cell(row=row, column=1,
                value=f"⚠ WARNING: calibration config differs from measurement: {'; '.join(warnings)}")
            warn_cell.font = Font(bold=True, color='9C0006')
            warn_cell.fill = PatternFill('solid', fgColor='FFC7CE')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 2

        # Index baseline/dut rows by frequency index
        base_by_f = {r['freq_idx']: r for r in baseline_rows}
        dut_by_f  = {r['freq_idx']: r for r in dut_rows}
        nominal_value = dut_meta.get('nominal_ohm')

        # === Table 5: calibrated results ===
        cal_headers = [
            'F idx', 'Frequency (Hz)', 'OSR',
            'R_cal (Ω)', 'X_cal (Ω)', '|Z|_cal (Ω)', 'Phase_cal (°)',
            'Δ nom abs (Ω)', 'Δ nom (%)'
        ]
        cal_nfmt = [fmt_int, fmt_int, fmt_int,
                    fmt_num, fmt_num, fmt_num, fmt_num,
                    fmt_num, fmt_pct]

        cal_rows = []
        all_freqs = sorted(set(base_by_f.keys()) & set(dut_by_f.keys()) & set(cal_k.keys()))
        for fi in all_freqs:
            base = base_by_f[fi]; dut = dut_by_f[fi]
            # Complex ADC after baseline subtraction
            adc_real = dut['real_adc_avg'] - base['real_adc_avg']
            adc_imag = dut['imag_adc_avg'] - base['imag_adc_avg']
            z_uncal = complex(adc_real, adc_imag) / K
            # Apply complex calibration factor k
            z_cal = cal_k[fi] * z_uncal
            mag_cal = abs(z_cal)
            phase_cal = math.degrees(cmath.phase(z_cal)) if mag_cal > 0 else 0

            delta_abs = ''
            delta_pct = ''
            if isinstance(nominal_value, (int, float)) and nominal_value != 0:
                delta_abs = mag_cal - nominal_value
                delta_pct = delta_abs / abs(nominal_value) * 100.0

            cal_rows.append([
                fi, dut['freq_hz'], dut['osr'],
                z_cal.real, z_cal.imag, mag_cal, phase_cal,
                delta_abs if isinstance(delta_abs, (int, float)) else '',
                delta_pct if isinstance(delta_pct, (int, float)) else '',
            ])

        ref_ohm = cal.get('reference_resistor_ohm', '?')
        t5_title = (f"Table 4 — Calibrated results "
                    f"(calibration: ref={ref_ohm}Ω, n_freq={len(cal_k)})")
        row = self._write_data_table(
            ws, row, t5_title, cal_headers, cal_rows,
            title_fill=title_fill, header_fill=header_fill, border=border,
            number_formats=cal_nfmt
        )

        # Highlight the |Z|_cal column (column F = 6)
        t5_header_row = row - len(cal_rows) - 1
        cal_col_fill = PatternFill('solid', fgColor='D5E3F4')
        cal_col_font = Font(bold=True, color='1F4E78')
        header_cell = ws.cell(row=t5_header_row, column=6)
        header_cell.fill = PatternFill('solid', fgColor='1F4E78')
        header_cell.font = Font(bold=True, color='FFFFFF')
        for r_idx in range(t5_header_row + 1, row):
            cell = ws.cell(row=r_idx, column=6)
            cell.fill = cal_col_fill
            cell.font = cal_col_font

        row += 2

        # === Table 6: before vs after calibration statistics ===
        stat_headers = [
            'OSR band', 'n_freq',
            'Uncal mean |Z| (Ω)', 'Uncal err mean (%)', 'Uncal err std (%)',
            'Cal mean |Z| (Ω)',   'Cal err mean (%)',   'Cal err std (%)',
        ]
        stat_nfmt = ['', fmt_int, fmt_num, fmt_pct, fmt_pct, fmt_num, fmt_pct, fmt_pct]

        bands = [
            ('F=0..30 (OSR=128)',   lambda fi: fi <= 30),
            ('F=31..51 (OSR=256)',  lambda fi: 31 <= fi <= 51),
            ('F=52 (OSR=512)',      lambda fi: fi == 52),
            ('F=53..59 (OSR=1024)', lambda fi: fi >= 53),
        ]

        import statistics as _stats

        def _mean_std(vals):
            if not vals: return (0.0, 0.0)
            return (_stats.mean(vals), _stats.stdev(vals) if len(vals) > 1 else 0.0)

        stat_rows = []
        for label, pred in bands:
            fis = [fi for fi in all_freqs if pred(fi)]
            if not fis:
                continue
            # korrigeerimata
            unc_mags, unc_errs = [], []
            cal_mags, cal_errs = [], []
            for fi in fis:
                base = base_by_f[fi]; dut = dut_by_f[fi]
                adc_r = dut['real_adc_avg'] - base['real_adc_avg']
                adc_i = dut['imag_adc_avg'] - base['imag_adc_avg']
                z_u = complex(adc_r, adc_i) / K
                z_c = cal_k[fi] * z_u
                um = abs(z_u); cm = abs(z_c)
                unc_mags.append(um); cal_mags.append(cm)
                if isinstance(nominal_value, (int, float)) and nominal_value != 0:
                    unc_errs.append((um - nominal_value) / nominal_value * 100.0)
                    cal_errs.append((cm - nominal_value) / nominal_value * 100.0)
            u_mean, _ = _mean_std(unc_mags)
            c_mean, _ = _mean_std(cal_mags)
            ue_mean, ue_std = _mean_std(unc_errs)
            ce_mean, ce_std = _mean_std(cal_errs)
            stat_rows.append([
                label, len(fis),
                u_mean,
                ue_mean if unc_errs else '',
                ue_std if unc_errs else '',
                c_mean,
                ce_mean if cal_errs else '',
                ce_std if cal_errs else '',
            ])

        row = self._write_data_table(
            ws, row, 'Table 5 — Statistics by OSR band (uncalibrated vs calibrated)',
            stat_headers, stat_rows,
            title_fill=title_fill, header_fill=header_fill, border=border,
            number_formats=stat_nfmt
        )

        if log_widget:
            log_widget.append(f"  ✓ Calibration applied "
                              f"(ref={ref_ohm}Ω, {len(cal_rows)} frequencies)")

        return row
# ──────────────────────────────────────────────────────────
# Tab 4 — Log / Terminal

# ──────────────────────────────────────────────────────────
class LogTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        top = QHBoxLayout()
        top.addWidget(QLabel("UART log / debug output"))
        top.addStretch()
        self.cb_scroll = QCheckBox("Auto scroll")
        self.cb_scroll.setChecked(True)
        top.addWidget(self.cb_scroll)
        btn_clear = QPushButton("Clear")
        btn_clear.clicked.connect(self._clear)
        top.addWidget(btn_clear)
        btn_save = QPushButton("Save")
        btn_save.clicked.connect(self._save)
        top.addWidget(btn_save)
        layout.addLayout(top)

        self.txt = QTextEdit()
        self.txt.setReadOnly(True)
        self.txt.setLineWrapMode(QTextEdit.NoWrap)
        layout.addWidget(self.txt)

        # Command line input + quick-send buttons
        cmd_row = QHBoxLayout()
        self.le_cmd = QLineEdit()
        self.le_cmd.setPlaceholderText("Send a command to STM32 (e.g. ? or F30 or S)...")
        self.le_cmd.returnPressed.connect(self._send_cmd)
        self.btn_send = QPushButton("Send")
        self.btn_send.clicked.connect(self._send_cmd)
        cmd_row.addWidget(self.le_cmd, 1)
        cmd_row.addWidget(self.btn_send)
        layout.addLayout(cmd_row)

        # Quick-send buttons for register debugging
        quick_row = QHBoxLayout()
        quick_row.setSpacing(4)
        quick_row.addWidget(QLabel("Quick commands:"))
        for label, cmd in [
            ("? Dump",      "?"),
            ("Q 0x20 CFG1", "Q20"),
            ("Q 0x21 CFG2", "Q21"),
            ("Q 0x22 CFG3", "Q22"),
            ("Q 0x24 CFG5", "Q24"),
            ("Q 0x25 CFG6", "Q25"),
            ("Q 0x17 PLL",  "Q17"),
            ("S Start",     "S"),
            ("P Stop",      "P"),
        ]:
            b = QPushButton(label)
            b.setMaximumHeight(26)
            b.clicked.connect(lambda _, c=cmd: self._send_raw(c))
            quick_row.addWidget(b)
        quick_row.addStretch()
        layout.addLayout(quick_row)

        self._port = ''
        self._baud = 1000000

    def set_port(self, port, baud):
        self._port = port
        self._baud = baud

    def append(self, text):
        ts = datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]
        colored = text
        # Color coding
        if text.startswith('[UART]'):
            color = COLORS['accent2']
        elif text.startswith('[STM32]') or text.startswith('[TULEMUS]'):
            color = COLORS['green']
        elif any(w in text for w in ('ERROR', 'VIGA', 'TIMEOUT')):
            color = COLORS['red']
        elif text.startswith('✓'):
            color = COLORS['green']
        elif text.startswith('⚠'):
            color = COLORS['yellow']
        else:
            color = COLORS['text']

        self.txt.append(
            f'<span style="color:{COLORS["text_dim"]}">{ts}</span>  '
            f'<span style="color:{color}">{text}</span>'
        )
        if self.cb_scroll.isChecked():
            self.txt.verticalScrollBar().setValue(
                self.txt.verticalScrollBar().maximum())

    def _clear(self):
        self.txt.clear()

    def _save(self):
        fname, _ = QFileDialog.getSaveFileName(
            self, "Save log", f"log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            "Tekstifailid (*.txt)")
        if fname:
            with open(fname, 'w') as f:
                f.write(self.txt.toPlainText())

    def _send_cmd(self):
        cmd = self.le_cmd.text().strip()
        if not cmd or not self._port:
            return
        try:
            s = serial.Serial(self._port, self._baud, timeout=2)
            s.write((cmd + '\n').encode())
            time.sleep(0.3)
            resp = []
            while s.in_waiting:
                l = s.readline().decode('ascii', errors='ignore').strip()
                if l:
                    resp.append(l)
            s.close()
            self.append(f"[UART] → {cmd}")
            for r in resp:
                self.append(f"[STM32] {r}")
        except Exception as e:
            self.append(f"ERROR sending command: {e}")
        self.le_cmd.clear()

    def _send_raw(self, cmd):
        """Send a command directly without using the command line (quick-send buttons)."""
        if not self._port:
            self.append("ERROR: no port selected. Go to the 'Overview' tab and pick a port.")
            return
        try:
            s = serial.Serial(self._port, self._baud, timeout=2)
            s.write((cmd + '\n').encode())
            time.sleep(0.3)
            resp = []
            while s.in_waiting:
                l = s.readline().decode('ascii', errors='ignore').strip()
                if l:
                    resp.append(l)
            s.close()
            self.append(f"[UART] → {cmd}")
            for r in resp:
                self.append(f"[STM32] {r}")
        except Exception as e:
            self.append(f"ERROR sending command: {e}")


# ──────────────────────────────────────────────────────────
# Main application window
# ──────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    """Top-level window.  Owns the single ScanWorker/QThread pair and routes its signals
    to the appropriate tabs.  All measurement start/stop logic lives here so that only
    one measurement can run at a time regardless of which tab initiated it."""
    def __init__(self):
        super().__init__()
        self.setWindowTitle("MAX30009 BioZ Juhtpaneel  v1.0")
        self.setMinimumSize(1000, 700)
        self.resize(1200, 780)
        self.setStyleSheet(STYLE_MAIN)

        self._worker = None
        self._thread = None
        self._current_cfg = {}

        self._build_ui()
        self._setup_status_bar()

        # Uuenda olek iga 5s
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(5000)

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ── Header bar ──
        header = QWidget()
        header.setFixedHeight(42)
        header.setStyleSheet(f"background: {COLORS['bg3']}; border-bottom: 1px solid {COLORS['border']};")
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(16, 0, 16, 0)

        title = QLabel("MAX30009  BioZ  Control Panel")
        title.setStyleSheet(f"color: {COLORS['accent2']}; font-size: 14px; font-weight: bold; letter-spacing: 2px;")
        subtitle = QLabel("Bioimpedance measurement system  |  STM32WB55  ↔  PC")
        subtitle.setStyleSheet(f"color: {COLORS['text_dim']}; font-size: 11px;")

        h_layout.addWidget(title)
        h_layout.addWidget(subtitle)
        h_layout.addStretch()

        main_layout.addWidget(header)

        # ── Vahekaardid ──
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.North)

        self.tab_dash = DashboardTab()
        self.tab_single = SingleMeasTab()
        self.tab_scan = ScanTab()
        self.tab_calib_builder = CalibrationTab()
        self.tab_log = LogTab()

        self.tabs.addTab(self.tab_dash,          "  📊  Overview  ")
        self.tabs.addTab(self.tab_single,        "  🔬  BIA  ")
        self.tabs.addTab(self.tab_scan,          "  📈  BIS  ")
        self.tabs.addTab(self.tab_calib_builder, "  🗂  Calibration Builder  ")
        self.tabs.addTab(self.tab_log,           "  🖥  Terminal / Log  ")

        main_layout.addWidget(self.tabs)

        # ── Signaalid ──
        self.tab_single.start_requested.connect(self._on_start)
        self.tab_scan.start_requested.connect(self._on_start)
        self.tab_single.btn_stop.clicked.connect(self._on_stop)
        self.tab_scan.btn_stop.clicked.connect(self._on_stop)

    def _setup_status_bar(self):
        self.status_bar = self.statusBar()
        self.lbl_sb_port = QLabel("Port: —")
        self.lbl_sb_time = QLabel("")
        self.lbl_sb_state = QLabel("Idle")
        self.status_bar.addWidget(self.lbl_sb_port)
        self.status_bar.addPermanentWidget(self.lbl_sb_state)
        self.status_bar.addPermanentWidget(self.lbl_sb_time)

    def _tick(self):
        self.lbl_sb_time.setText(datetime.datetime.now().strftime('%H:%M:%S  '))

    def _on_start(self, cfg):
        """Create and start a ScanWorker for the requested measurement.

        cfg dict keys:
          mode        — 'single' | 'scan'
          stim_idx    — stimulus current index (0..15)
          gain_idx    — PGA gain index (0..3)
          ahpf/dhpf/dlpf/dcrestore/extcap/amprange/ampbw — advanced filter/amp settings
          component   — label string for CSV/Excel
          freq_start/freq_end — scan range (scan mode only)
          single_freq — fixed frequency index (single mode only)
          _calib_phase — 'baseline' | 'dut' | None  (injected by ScanTab for calibrated mode)
        """
        if self._thread and self._thread.isRunning():
            return

        port = self.tab_dash.get_port()
        baud = self.tab_dash.get_baud()

        if not port or port.startswith('('):
            self.tab_log.append("ERROR: Select the connection port on the Overview tab!")
            self.tabs.setCurrentIndex(0)
            return

        self._current_cfg = cfg
        self.tab_log.set_port(port, baud)
        self.tab_log.append(f"── Starting: {cfg['mode']} | {port} @ {baud} ──")
        self.lbl_sb_port.setText(f"Port: {port} @ {baud}")

        self._worker = ScanWorker()
        self._worker.port = port
        self._worker.baud = baud
        self._worker.stim_idx = cfg['stim_idx']
        self._worker.gain_idx = cfg['gain_idx']
        self._worker.component = cfg['component']
        self._worker.mode      = cfg['mode']
        self._worker.ahpf      = cfg.get('ahpf', 15)
        self._worker.dhpf      = cfg.get('dhpf', 0)
        self._worker.dlpf      = cfg.get('dlpf', 0)
        self._worker.dcrestore = cfg.get('dcrestore', 0)
        self._worker.extcap    = cfg.get('extcap', 0)
        self._worker.amprange  = cfg.get('amprange', 3)
        self._worker.ampbw     = cfg.get('ampbw', 3)

        if cfg['mode'] == 'single':
            self._worker.single_freq = cfg['single_freq']
            self._worker.log_mode    = cfg.get('log_mode', 'averaged')
            self._worker.single_result.connect(self._on_single_result)
            self._worker.flags_result.connect(self.tab_single.update_flags)
            # raw_pair salvestatakse SingleTab-is, et hiljem avatud aken saaks ka andmeid
            self.tab_single.set_worker(self._worker)
        else:
            self._worker.freq_start = cfg['freq_start']
            self._worker.freq_end   = cfg['freq_end']
            self._worker.freq_done.connect(self._on_freq_done)
            self._worker.progress.connect(self._on_progress)
            self.tab_scan.scan_meta = {
                'component':   cfg['component'],
                'nominal_ohm': cfg.get('nominal_ohm'),
                'stim_idx':    cfg['stim_idx'],
                'gain_idx':    cfg['gain_idx'],
                'freq_start':  cfg['freq_start'],
                'freq_end':    cfg['freq_end'],
                'ahpf':        cfg.get('ahpf', 15),
                'dhpf':        cfg.get('dhpf', 0),
                'dlpf':        cfg.get('dlpf', 0),
                'dcrestore':   cfg.get('dcrestore', 0),
                'extcap':      cfg.get('extcap', 0),
                'amprange':    cfg.get('amprange', 3),
                'ampbw':       cfg.get('ampbw', 3),
                'calib_phase': cfg.get('_calib_phase'),
                'timestamp':   datetime.datetime.now().isoformat(),
            }

        self._worker.log.connect(self.tab_log.append)
        self._worker.finished.connect(self._on_finished)

        self._thread = QThread()
        self._worker.moveToThread(self._thread)

        if cfg['mode'] == 'single':
            self._thread.started.connect(self._worker.run_single)
        else:
            self._thread.started.connect(self._worker.run_scan)

        self._thread.start()

        # UI olek
        if cfg['mode'] == 'single':
            self.tab_single.set_running(True)
            self.lbl_sb_state.setText("⏳ Measurement...")
        else:
            self.tab_scan.set_running(True)
            self.lbl_sb_state.setText("⏳ Scan...")

    def _on_stop(self):
        if self._worker:
            self._worker.stop()
            self.tab_log.append("⚠ Stopping...")

    def _on_single_result(self, R, X, Z, phase, avg_r, avg_i):
        freq_idx = self._current_cfg.get('single_freq', 30)
        comp = self._current_cfg.get('component', '?')
        self.tab_single.update_result(R, X, Z, phase, comp, freq_idx, avg_r, avg_i)
        self.tab_dash.update_single_result(R, X, Z, phase)

    def _on_freq_done(self, freq_idx, n_samples, R, X):
        self.tab_scan.add_freq_result(freq_idx, n_samples, R, X)
        Z = math.sqrt(R*R + X*X)
        phase = math.degrees(math.atan2(X, R)) if Z > 0 else 0
        self.tab_dash.update_single_result(R, X, Z, phase)
        # Uuenda dashboard info
        if freq_idx < len(FREQ_HZ):
            self.tab_dash.lbl_freq_info.setText(
                f"{format_freq(FREQ_HZ[freq_idx])}  (F={freq_idx})")
        self.tab_dash.lbl_current_info.setText(
            format_current(self._current_cfg.get('stim_idx', 12)))
        self.tab_dash.lbl_gain_info.setText(
            f"{GAINS[self._current_cfg.get('gain_idx', 0)]:.0f}x")
        self.tab_dash.lbl_samples_info.setText(str(n_samples))

    def _on_progress(self, done, total):
        self.tab_scan.update_progress(done, total)

    def _on_finished(self, ok, msg):
        self.tab_log.append(f"{'✓' if ok else '✗'} Stopped: {msg}")
        if self._worker and hasattr(self._worker, 'raw_rows_buf') and self._worker.raw_rows_buf:
            self.tab_scan.raw_rows.extend(self._worker.raw_rows_buf)
            self.tab_log.append(f"  {len(self._worker.raw_rows_buf)} raw ADC rows stored in memory")
        self.tab_single.set_running(False)
        self.lbl_sb_state.setText("✓ Ready" if ok else "✗ Error")

        scan_tab = self.tab_scan
        calib_phase = self._current_cfg.get('_calib_phase')

        # ── Calibrated mode: baseline finished → keep in memory, trigger DUT scan ──
        if ok and calib_phase == 'baseline' and scan_tab.raw_rows:
            scan_tab.baseline_raw_rows = list(scan_tab.raw_rows)
            scan_tab.baseline_meta = dict(scan_tab.scan_meta)
            scan_tab.baseline_avg = scan_tab._compute_avg_from_raw(scan_tab.raw_rows)
            self.tab_log.append(f"  ✓ Baseline (0Ω) stored in memory "
                                f"({len(scan_tab.raw_rows)} rows).")

            try:
                scan_tab.baseline_csv_path = scan_tab._save_raw_csv(
                    scan_tab.baseline_raw_rows,
                    scan_tab.baseline_meta,
                    scan_tab.output_folder,
                    label='0ohm',
                    log_widget=self.tab_log,
                )
            except Exception as e:
                self.tab_log.append(f"  ✗ Baseline raw CSV save failed: {e}")

            self.tab_log.append("  → Requesting DUT now...")

            if self._thread:
                self._thread.quit()
                self._thread.wait()

            scan_tab.start_dut_phase()
            return

        # ── Calibrated mode: DUT finished → subtract baseline, write Excel ──
        if ok and calib_phase == 'dut' and scan_tab.raw_rows:
            dut_raw = list(scan_tab.raw_rows)
            dut_meta = dict(scan_tab.scan_meta)
            self.tab_log.append(f"  ✓ DUT ({dut_meta.get('nominal_ohm')}Ω) "
                                f"stored in memory ({len(dut_raw)} rows).")

            try:
                scan_tab.dut_csv_path = scan_tab._save_raw_csv(
                    dut_raw,
                    dut_meta,
                    scan_tab.output_folder,
                    label=f"{dut_meta.get('nominal_ohm')}ohm",
                    log_widget=self.tab_log,
                )
            except Exception as e:
                self.tab_log.append(f"  ✗ DUT raw CSV save failed: {e}")

            try:
                xlsx_path, sheet_name = scan_tab._update_analysis_xlsx(
                    baseline_raw=scan_tab.baseline_raw_rows,
                    baseline_meta=scan_tab.baseline_meta,
                    dut_raw=dut_raw,
                    dut_meta=dut_meta,
                    log_widget=self.tab_log)
                msg_lines = [
                    'Measurement saved:',
                    '',
                    f"0Ω raw CSV: {scan_tab.baseline_csv_path or '-'}",
                    f"DUT raw CSV: {scan_tab.dut_csv_path or '-'}",
                    f"Excel: {xlsx_path}",
                    f"Sheet: {sheet_name}",
                ]
                QMessageBox.information(scan_tab, 'Ready', '\n'.join(msg_lines))
            except ImportError as e:
                err_msg = (f"openpyxl is not installed!\n\n"
                           f"Open a terminal and run:\n"
                           f"    pip install openpyxl\n\n"
                           f"Detail: {e}")
                self.tab_log.append(f"  ✗ ANALYSIS ERROR: {err_msg}")
                QMessageBox.critical(scan_tab, 'Analysis error', err_msg)
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                self.tab_log.append(f"  ✗ ANALYSIS ERROR: {e}")
                for ln in tb.splitlines():
                    self.tab_log.append(f"    {ln}")
                QMessageBox.critical(scan_tab, 'Analysis error',
                    f"Analysis failed:\n\n{e}\n\nSee details on the Terminal / Log tab.")

            scan_tab.calib_mode = False
            scan_tab.calib_phase = None
            scan_tab.pending_cfg = None
            scan_tab.set_running(False)

            if self._thread:
                self._thread.quit()
                self._thread.wait()
            return

        # ── Tavarežiim või katkestus ──
        scan_tab.set_running(False)

        # CSV export after scan — asks for folder (non-calibrated mode)
        if ok and self._worker and self._worker.mode == 'scan' \
                and scan_tab.raw_rows and not calib_phase:
            self.tab_log.append("  → Select output folder...")
            scan_tab._auto_export(self.tab_log)

        # Calibrated mode abort → reset state
        if not ok and calib_phase:
            scan_tab.calib_mode = False
            scan_tab.calib_phase = None
            scan_tab.pending_cfg = None

        if self._thread:
            self._thread.quit()
            self._thread.wait()


# ──────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────
def main():
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    app.setApplicationName("MAX30009 BioZ Juhtpaneel")

    # HiDPI
    app.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    app.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
